"""Stanvard School ERP - FastAPI backend."""
import os
import logging
import hmac
import hashlib
import io
import csv
import re
import asyncio
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Any
from pathlib import Path

from fastapi import FastAPI, APIRouter, Depends, HTTPException, Request, Query, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
import razorpay

from database import (
    db, mongo_client,
    schools_col, users_col, students_col, classes_col,
    fee_heads_col, fee_plans_col, fee_assignments_col,
    payments_col, receipts_col, razorpay_orders_col,
    attendance_col, homework_col, timetable_col,
    events_col, circulars_col, gallery_col, staff_col,
    notifications_col, audit_col, settings_col,
    discount_approvals_col, exams_col, marks_col, message_logs_col,
    get_next_sequence,
)
from models import (
    School, SchoolCreate, SchoolUpdate,
    User, UserCreate, UserUpdate, LoginRequest, LoginResponse,
    Student, StudentCreate, StudentUpdate,
    ClassRoom, ClassCreate,
    FeeHead, FeeHeadCreate, FeePlan, FeePlanCreate, FeeAssignment, FeeAssignmentCreate,
    FeeAssignmentUpdate, FeeAssignmentItem, FeeAssignmentInstallment,
    Payment, PaymentCreate, PaymentLineItem, PaymentEdit, PaymentVoid,
    DiscountApproval, DiscountApprovalReview, DiscountApprovalCollect,
    RazorpayOrderRequest, RazorpayVerifyRequest,
    AttendanceRecord, AttendanceBulkMark,
    Homework, HomeworkCreate,
    Timetable, TimetableCreate,
    Event, EventCreate,
    Circular, CircularCreate,
    GalleryAlbum, GalleryAlbumCreate,
    Staff, StaffCreate,
    Notification, NotificationCreate,
    SchoolSettings,
    Exam, ExamCreate, ExamUpdate, MarksBulkSave, FeeReminderRequest,
    now_iso, new_id,
)
from auth import (
    hash_password, verify_password, create_access_token,
    get_current_user, require_roles, current_school_id, resolve_school_id, resolve_school_id_safe,
)
from audit import log_audit
from pdf_utils import generate_receipt_pdf, generate_report_pdf

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

RAZORPAY_KEY_ID = os.environ.get('RAZORPAY_KEY_ID', '')
RAZORPAY_KEY_SECRET = os.environ.get('RAZORPAY_KEY_SECRET', '')
RAZORPAY_WEBHOOK_SECRET = os.environ.get('RAZORPAY_WEBHOOK_SECRET', '')

rzp_client = None
if RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET:
    rzp_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))

# When no real Razorpay keys are configured we run in MOCK/test mode: orders are
# generated locally and signatures are computed against MOCK_KEY_SECRET so the
# full order-create -> verify flow can be exercised end-to-end without hitting
# Razorpay servers (no real capture). This mirrors sandbox behaviour for QA.
RAZORPAY_MOCK = rzp_client is None
MOCK_KEY_ID = os.environ.get('MOCK_KEY_ID', 'rzp_test_mock')
MOCK_KEY_SECRET = os.environ.get('MOCK_KEY_SECRET', 'mock_secret_stanvard')

# Fee collection calendar: annual fees are split across 10 months only, skipping
# June (6) and March (3). Collection months are Apr, May, Jul, Aug, Sep, Oct, Nov,
# Dec, Jan, Feb. June & March are shown as "no fee" months in the schedule.
FEE_EXCLUDED_MONTHS = {6, 3}
FEE_COLLECTION_MONTH_COUNT = 12 - len(FEE_EXCLUDED_MONTHS)  # 10

app = FastAPI(title='Stanvard School ERP API', version='1.0.0')


@app.get('/health')
async def root_health():
    return {'ok': True, 'status': 'healthy'}


api = APIRouter(prefix='/api')

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger('stanvard')

# ----- utils -----
def strip_password(u: dict) -> dict:
    d = dict(u or {})
    d.pop('password_hash', None)
    d.pop('_id', None)
    return d


def clean(doc):
    if doc is None:
        return None
    if isinstance(doc, list):
        return [clean(d) for d in doc]
    if isinstance(doc, dict):
        return {k: v for k, v in doc.items() if k != '_id'}
    return doc


# =====================================================
# HEALTH / ROOT
# =====================================================
@api.get('/')
async def root():
    return {'app': 'Stanvard School ERP', 'status': 'running'}


@api.get('/health')
async def health():
    return {'ok': True, 'time': now_iso()}


# =====================================================
# AUTH
# =====================================================
def parent_linked_student_ids(user: Dict[str, Any]) -> List[str]:
    """Return the union of a parent user's linked student IDs.

    Merges the legacy single-child field `linked_student_id` with the
    multi-child list `linked_student_ids` so that both new and old accounts
    are supported without migration.
    """
    ids = list(user.get('linked_student_ids') or [])
    legacy = user.get('linked_student_id')
    if legacy and legacy not in ids:
        ids.append(legacy)
    return [i for i in ids if i]


def parent_can_access_student(user: Dict[str, Any], student_id: Optional[str]) -> bool:
    if not student_id:
        return False
    return student_id in parent_linked_student_ids(user)


@api.post('/auth/login', response_model=LoginResponse)
async def login(body: LoginRequest, request: Request):
    identifier = (body.email or '').strip()
    if not identifier:
        raise HTTPException(status_code=400, detail='Email or mobile number is required')

    # 1. Try email match (case-insensitive)
    user = await users_col.find_one({'email': identifier.lower()}, {'_id': 0})

    # 2. If not found and identifier looks like a phone number, try phone match.
    #    We accept +91, spaces, dashes, etc. and try both the full-digit string
    #    and the last 10 digits (Indian mobile) against the stored phone.
    if not user:
        digits_only = ''.join(ch for ch in identifier if ch.isdigit())
        candidates: List[str] = []
        if digits_only and len(digits_only) >= 7:
            candidates.append(digits_only)
            if len(digits_only) > 10:
                candidates.append(digits_only[-10:])
        for cand in candidates:
            # Exact match, then "ends-with" match on stored phone
            user = await users_col.find_one({'phone': cand}, {'_id': 0})
            if not user:
                user = await users_col.find_one(
                    {'phone': {'$regex': f'{cand}$'}}, {'_id': 0}
                )
            if user:
                break

    if not user or not verify_password(body.password, user.get('password_hash', '')):
        raise HTTPException(status_code=401, detail='Invalid credentials')
    if user.get('status') != 'active':
        raise HTTPException(status_code=403, detail='Account is inactive')
    token = create_access_token({'sub': user['id'], 'role': user['role']})
    await log_audit(action='auth.login', current_user=user,
                    ip_address=request.client.host if request.client else None)
    return LoginResponse(access_token=token, user=strip_password(user))


@api.get('/auth/me')
async def me(current=Depends(get_current_user)):
    return strip_password(current)


@api.get('/auth/my-schools')
async def my_schools(current=Depends(get_current_user)):
    """Schools accessible to current user."""
    if current['role'] == 'super_admin':
        rows = await schools_col.find({'status': {'$ne': 'deleted'}}, {'_id': 0}).to_list(1000)
    else:
        sid = current.get('school_id')
        rows = await schools_col.find({'id': sid}, {'_id': 0}).to_list(1) if sid else []
    return rows


# =====================================================
# SCHOOLS (super admin)
# =====================================================
@api.get('/schools')
async def list_schools(current=Depends(get_current_user)):
    if current['role'] == 'super_admin':
        rows = await schools_col.find({'status': {'$ne': 'deleted'}}, {'_id': 0}).to_list(1000)
    else:
        sid = current.get('school_id')
        rows = await schools_col.find({'id': sid}, {'_id': 0}).to_list(1) if sid else []
    return rows


@api.post('/schools', dependencies=[Depends(require_roles('super_admin'))])
async def create_school(body: SchoolCreate, current=Depends(get_current_user)):
    s = School(**body.model_dump())
    doc = s.model_dump()
    await schools_col.insert_one(doc)
    await log_audit(action='school.create', current_user=current,
                    entity_type='school', entity_id=s.id, details={'name': s.name})
    return clean(doc)


@api.patch('/schools/{school_id}', dependencies=[Depends(require_roles('super_admin'))])
async def update_school(school_id: str, body: SchoolUpdate, current=Depends(get_current_user)):
    upd = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    upd['updated_at'] = now_iso()
    r = await schools_col.update_one({'id': school_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'School not found')
    await log_audit(action='school.update', current_user=current,
                    entity_type='school', entity_id=school_id, details=upd)
    doc = await schools_col.find_one({'id': school_id}, {'_id': 0})
    return clean(doc)


@api.delete('/schools/{school_id}', dependencies=[Depends(require_roles('super_admin'))])
async def archive_school(school_id: str, current=Depends(get_current_user)):
    r = await schools_col.update_one({'id': school_id},
                                     {'$set': {'status': 'archived', 'updated_at': now_iso()}})
    if not r.matched_count:
        raise HTTPException(404, 'School not found')
    await log_audit(action='school.archive', current_user=current,
                    entity_type='school', entity_id=school_id)
    return {'ok': True}


# =====================================================
# USERS
# =====================================================
@api.get('/users')
async def list_users(current=Depends(get_current_user),
                    role: Optional[str] = None,
                    school_id: Optional[str] = None):
    q: Dict[str, Any] = {}
    if role:
        q['role'] = role
    if current['role'] == 'super_admin':
        if school_id:
            q['school_id'] = school_id
    else:
        q['school_id'] = current['school_id']
    rows = await users_col.find(q, {'_id': 0}).to_list(2000)
    return [strip_password(r) for r in rows]


@api.post('/users', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_user(body: UserCreate, current=Depends(get_current_user)):
    if body.role not in {'super_admin', 'school_admin', 'accountant', 'teacher', 'parent'}:
        raise HTTPException(400, 'Invalid role')
    if current['role'] == 'school_admin':
        # school admin cannot create super_admin, and must scope to own school
        if body.role == 'super_admin':
            raise HTTPException(403, 'Cannot create super admin')
        body.school_id = current['school_id']
    exists = await users_col.find_one({'email': body.email.lower()}, {'_id': 0})
    if exists:
        raise HTTPException(400, 'Email already registered')
    u = User(
        email=body.email.lower(),
        password_hash=hash_password(body.password),
        full_name=body.full_name,
        role=body.role,
        school_id=body.school_id,
        phone=body.phone,
        linked_student_id=body.linked_student_id,
        linked_student_ids=body.linked_student_ids or [],
        linked_class_ids=body.linked_class_ids,
    )
    await users_col.insert_one(u.model_dump())
    await log_audit(action='user.create', current_user=current,
                    entity_type='user', entity_id=u.id,
                    details={'email': u.email, 'role': u.role})
    return strip_password(u.model_dump())


@api.patch('/users/{user_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_user(user_id: str, body: UserUpdate, current=Depends(get_current_user)):
    upd: Dict[str, Any] = {}
    for k, v in body.model_dump(exclude_none=True).items():
        if k == 'password':
            upd['password_hash'] = hash_password(v)
        else:
            upd[k] = v
    upd['updated_at'] = now_iso()
    r = await users_col.update_one({'id': user_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'User not found')
    await log_audit(action='user.update', current_user=current,
                    entity_type='user', entity_id=user_id, details={k: v for k, v in upd.items() if k != 'password_hash'})
    doc = await users_col.find_one({'id': user_id}, {'_id': 0})
    return strip_password(doc)


# =====================================================
# STUDENTS
# =====================================================
@api.get('/students')
async def list_students(request: Request,
                        current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        class_id: Optional[str] = None,
                        section: Optional[str] = None,
                        status: Optional[str] = None,
                        search: Optional[str] = None,
                        limit: int = 500):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    # For parent role, restrict to their linked children
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        if not child_ids:
            return []
        q['id'] = {'$in': child_ids}
    if class_id:
        q['class_id'] = class_id
    if section:
        q['section'] = section
    if status:
        q['status'] = status
    if search:
        q['$or'] = [
            {'full_name': {'$regex': search, '$options': 'i'}},
            {'admission_number': {'$regex': search, '$options': 'i'}},
            {'father_name': {'$regex': search, '$options': 'i'}},
            {'phone': {'$regex': search, '$options': 'i'}},
        ]
    rows = await students_col.find(q, {'_id': 0}).limit(limit).to_list(limit)
    return rows


@api.get('/students/{student_id}')
async def get_student(student_id: str, request: Request, current=Depends(get_current_user)):
    s = await students_col.find_one({'id': student_id}, {'_id': 0})
    if not s:
        raise HTTPException(404, 'Student not found')
    # scope check
    if current['role'] != 'super_admin' and s['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')
    return s


@api.post('/students', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_student(body: StudentCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    adm_no = body.admission_number
    if not adm_no:
        seq = await get_next_sequence(f'adm_{sid}')
        school = await schools_col.find_one({'id': sid}, {'_id': 0})
        code = (school or {}).get('code', 'STV')
        adm_no = f'{code}-2025-{seq:04d}'
    data = body.model_dump(exclude_none=True)
    data.pop('school_id', None)
    student = Student(school_id=sid, admission_number=adm_no, **{k: v for k, v in data.items() if k != 'admission_number'})
    await students_col.insert_one(student.model_dump())
    await log_audit(action='student.create', current_user=current, school_id=sid,
                    entity_type='student', entity_id=student.id,
                    details={'admission_number': adm_no, 'name': student.full_name})
    return student.model_dump()


@api.patch('/students/{student_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_student(student_id: str, body: StudentUpdate, current=Depends(get_current_user)):
    upd = body.model_dump(exclude_none=True)
    upd['updated_at'] = now_iso()
    r = await students_col.update_one({'id': student_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'Student not found')
    await log_audit(action='student.update', current_user=current,
                    entity_type='student', entity_id=student_id, details=upd)
    return await students_col.find_one({'id': student_id}, {'_id': 0})


@api.delete('/students/{student_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_student(student_id: str, current=Depends(get_current_user)):
    r = await students_col.update_one({'id': student_id},
                                      {'$set': {'status': 'inactive', 'updated_at': now_iso()}})
    if not r.matched_count:
        raise HTTPException(404, 'Student not found')
    await log_audit(action='student.delete', current_user=current,
                    entity_type='student', entity_id=student_id)
    return {'ok': True}


# =====================================================
# CLASSES
# =====================================================
@api.get('/classes')
async def list_classes(request: Request, current=Depends(get_current_user),
                       school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    rows = await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(1000)
    return rows


@api.post('/classes', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_class(body: ClassCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    c = ClassRoom(school_id=sid, name=body.name, sections=body.sections, teacher_id=body.teacher_id)
    await classes_col.insert_one(c.model_dump())
    await log_audit(action='class.create', current_user=current, school_id=sid,
                    entity_type='class', entity_id=c.id, details={'name': c.name})
    return c.model_dump()


# =====================================================
# FEE HEADS
# =====================================================
@api.get('/fees/heads')
async def list_fee_heads(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await fee_heads_col.find({'school_id': sid}, {'_id': 0}).to_list(500)


@api.post('/fees/heads', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_fee_head(body: FeeHeadCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    fh = FeeHead(school_id=sid, name=body.name, category=body.category)
    await fee_heads_col.insert_one(fh.model_dump())
    await log_audit(action='fee_head.create', current_user=current, school_id=sid,
                    entity_type='fee_head', entity_id=fh.id, details={'name': fh.name})
    return fh.model_dump()


@api.patch('/fees/heads/{head_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_fee_head(head_id: str, body: FeeHeadCreate, current=Depends(get_current_user)):
    upd = body.model_dump(exclude_none=True)
    upd.pop('school_id', None)
    upd['updated_at'] = now_iso()
    r = await fee_heads_col.update_one({'id': head_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'Fee head not found')
    await log_audit(action='fee_head.update', current_user=current,
                    entity_type='fee_head', entity_id=head_id, details=upd)
    return await fee_heads_col.find_one({'id': head_id}, {'_id': 0})


@api.delete('/fees/heads/{head_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_fee_head(head_id: str, current=Depends(get_current_user)):
    head = await fee_heads_col.find_one({'id': head_id}, {'_id': 0})
    if not head:
        raise HTTPException(404, 'Fee head not found')
    # Safety: refuse delete if this head is used in any plan or assignment.
    plan_uses = await fee_plans_col.count_documents({'items.fee_head_id': head_id})
    assign_uses = await fee_assignments_col.count_documents({'custom_items.fee_head_id': head_id})
    if plan_uses or assign_uses:
        raise HTTPException(400, f'Cannot delete: fee head is used in {plan_uses} plan(s) and {assign_uses} assignment(s). Remove references first or deactivate the head instead.')
    await fee_heads_col.delete_one({'id': head_id})
    await log_audit(action='fee_head.delete', current_user=current,
                    entity_type='fee_head', entity_id=head_id, details={'name': head.get('name')})
    return {'ok': True}


# =====================================================
# FEE PLANS
# =====================================================
@api.get('/fees/plans')
async def list_fee_plans(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await fee_plans_col.find({'school_id': sid}, {'_id': 0}).to_list(500)


def _build_plan_installments(plan: dict, session: str) -> list:
    """Build the 10-month installment timeline for a plan (June & March skipped).
    Uses per-month overrides (plan.month_amounts) when set, else the plan's
    discounted equal split across the 10 collection months.

    One-time items (frequency == 'one_time') are NOT divided across months —
    each is charged in full in its chosen month (item.one_time_month), falling
    back to the first collection month. Only recurring items feed the equal
    monthly split."""
    months = _session_months(session)
    coll_month_nums = [m['month'] for m in months if m['month'] not in FEE_EXCLUDED_MONTHS]
    all_items = plan.get('items') or []
    recurring_items = [it for it in all_items if (it.get('frequency') or 'monthly') != 'one_time']
    one_time_items = [it for it in all_items if (it.get('frequency') or 'monthly') == 'one_time']
    recurring_gross = sum(float(it.get('amount') or 0) for it in recurring_items)
    bd = compute_plan_discount_breakdown(plan, recurring_gross, coll_month_nums)
    net_by_month = {int(r['month']): float(r['net']) for r in bd['per_month']}
    # One-time charges by target month (default: first collection month).
    first_coll = coll_month_nums[0] if coll_month_nums else None
    one_time_by_month: Dict[int, float] = {}
    for it in one_time_items:
        tm = it.get('one_time_month')
        try:
            tm = int(tm) if tm is not None else None
        except Exception:
            tm = None
        if tm not in coll_month_nums:
            tm = first_coll
        if tm is not None:
            one_time_by_month[tm] = one_time_by_month.get(tm, 0.0) + float(it.get('amount') or 0)
    override_by_month = {}
    for o in (plan.get('month_amounts') or []):
        try:
            override_by_month[int(o.get('month'))] = round(float(o.get('amount') or 0), 2)
        except Exception:
            continue
    installments = []
    for m in months:
        one_time_here = one_time_by_month.get(m['month'], 0.0)
        has_override = m['month'] in override_by_month
        if m['month'] in FEE_EXCLUDED_MONTHS and one_time_here <= 0 and not has_override:
            installments.append({'month': m['month'], 'year': m['year'], 'amount': 0.0,
                                 'status': 'skip', 'label': 'No Fee',
                                 'due_date': f"{m['year']}-{m['month']:02d}-15"})
        else:
            amt = override_by_month.get(m['month'], net_by_month.get(m['month'], 0.0))
            amt = round(float(amt) + one_time_here, 2)
            if amt <= 0:
                # Explicit zero override (or zero computed amount) = month skipped.
                installments.append({'month': m['month'], 'year': m['year'], 'amount': 0.0,
                                     'status': 'skip', 'label': 'No Fee',
                                     'due_date': f"{m['year']}-{m['month']:02d}-15"})
            else:
                installments.append({'month': m['month'], 'year': m['year'],
                                     'amount': amt, 'status': 'active',
                                     'due_date': f"{m['year']}-{m['month']:02d}-15"})
    return installments


async def _sync_plan_assignments(plan: dict, current: dict) -> int:
    """Auto-assign a class-wide plan to every active student of its class.
    Creates a fee assignment for students who have none this session, and for
    students who already have one, repoints it at THIS plan (replacing custom
    items) and applies the plan's 10-month installment timeline. Each student's
    own discount/concession on the assignment is preserved."""
    if not plan.get('class_id'):
        return 0
    session = plan.get('academic_session') or '2026-27'
    installments = _build_plan_installments(plan, session)
    students = await students_col.find(
        {'school_id': plan['school_id'], 'class_id': plan['class_id'], 'status': 'active'},
        {'_id': 0, 'id': 1},
    ).to_list(10000)
    count = 0
    for s in students:
        existing = await fee_assignments_col.find_one(
            {'student_id': s['id'], 'academic_session': session}, {'_id': 0})
        if existing:
            # Point the student's assignment at this class plan and apply the
            # plan's 10-month installment timeline. custom per-student items are
            # replaced by the plan; the student's own discount/concession is kept.
            await fee_assignments_col.update_one(
                {'id': existing['id']},
                {'$set': {'fee_plan_id': plan['id'], 'custom_items': [],
                          'installments': installments, 'updated_at': now_iso()}})
            count += 1
            continue
        a = FeeAssignment(
            school_id=plan['school_id'], student_id=s['id'], fee_plan_id=plan['id'],
            academic_session=session,
            installments=[FeeAssignmentInstallment(**i) for i in installments],
        )
        await fee_assignments_col.insert_one(a.model_dump())
        count += 1
    return count


async def _demote_sibling_plans(plan: dict) -> None:
    """Mark every OTHER plan of the same class+session as draft, so only the
    given plan stays live for that class."""
    if not plan.get('class_id'):
        return
    await fee_plans_col.update_many(
        {'school_id': plan['school_id'], 'class_id': plan['class_id'],
         'academic_session': plan.get('academic_session'), 'id': {'$ne': plan['id']}},
        {'$set': {'status': 'draft', 'is_active': False, 'updated_at': now_iso()}})


@api.post('/fees/plans', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_fee_plan(body: FeePlanCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    data = body.model_dump()
    _validate_plan_discounts(data)
    data['school_id'] = sid
    # LIVE/DRAFT: a class may hold many plans but only ONE is live at a time and
    # only the live plan auto-assigns to students. The first plan for a
    # class+session becomes live automatically; extra plans are drafts unless
    # the caller explicitly asks for live.
    requested_status = data.pop('status', None)
    existing_live = None
    if data.get('class_id'):
        existing_live = await fee_plans_col.find_one(
            {'school_id': sid, 'class_id': data['class_id'],
             'academic_session': data.get('academic_session'), 'status': 'live'}, {'_id': 0})
    make_live = (requested_status == 'live') or (existing_live is None and bool(data.get('class_id')))
    data['status'] = 'live' if make_live else 'draft'
    data['is_active'] = make_live
    p = FeePlan(**data)
    await fee_plans_col.insert_one(p.model_dump())
    assigned = 0
    if make_live:
        await _demote_sibling_plans(p.model_dump())
        assigned = await _sync_plan_assignments(p.model_dump(), current)
    await log_audit(action='fee_plan.create', current_user=current, school_id=sid,
                    entity_type='fee_plan', entity_id=p.id,
                    details={'name': p.name, 'status': p.status, 'assigned_students': assigned})
    out = p.model_dump()
    out['assigned_students'] = assigned
    return out


@api.patch('/fees/plans/{plan_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_fee_plan(plan_id: str, body: FeePlanCreate, current=Depends(get_current_user)):
    upd = body.model_dump(exclude_none=True)
    _validate_plan_discounts(upd)
    upd.pop('school_id', None)
    # status is managed via /set-live, never through a generic edit.
    upd.pop('status', None)
    upd['updated_at'] = now_iso()
    r = await fee_plans_col.update_one({'id': plan_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'Plan not found')
    doc = await fee_plans_col.find_one({'id': plan_id}, {'_id': 0})
    # Only a LIVE plan pushes its edits down to student assignments. Editing a
    # draft plan never touches student schedules.
    assigned = 0
    if doc and doc.get('status') == 'live':
        assigned = await _sync_plan_assignments(doc, current)
    await log_audit(action='fee_plan.update', current_user=current,
                    entity_type='fee_plan', entity_id=plan_id,
                    details={**upd, 'assigned_students': assigned})
    if doc:
        doc['assigned_students'] = assigned
    return doc


@api.post('/fees/plans/{plan_id}/set-live', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def set_plan_live(plan_id: str, current=Depends(get_current_user)):
    """Make this plan the LIVE one for its class+session (demoting the previous
    live plan to draft) and auto-assign it to every active student of the class."""
    plan = await fee_plans_col.find_one({'id': plan_id}, {'_id': 0})
    if not plan:
        raise HTTPException(404, 'Plan not found')
    if not plan.get('class_id'):
        raise HTTPException(400, 'Only class-wide plans can be set live.')
    await fee_plans_col.update_one({'id': plan_id},
                                   {'$set': {'status': 'live', 'is_active': True, 'updated_at': now_iso()}})
    plan['status'] = 'live'
    await _demote_sibling_plans(plan)
    assigned = await _sync_plan_assignments(plan, current)
    await log_audit(action='fee_plan.set_live', current_user=current, school_id=plan['school_id'],
                    entity_type='fee_plan', entity_id=plan_id,
                    details={'name': plan.get('name'), 'assigned_students': assigned})
    return {'ok': True, 'status': 'live', 'assigned_students': assigned}


@api.post('/fees/plans/{plan_id}/duplicate', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def duplicate_fee_plan(plan_id: str, current=Depends(get_current_user)):
    """Create a DRAFT copy of a plan. Never touches the live plan or students."""
    plan = await fee_plans_col.find_one({'id': plan_id}, {'_id': 0})
    if not plan:
        raise HTTPException(404, 'Plan not found')
    if current['role'] != 'super_admin' and plan['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    data = {k: v for k, v in plan.items() if k not in ('id', 'created_at', 'updated_at')}
    data['name'] = f"{plan.get('name', 'Plan')} (Copy)"
    data['status'] = 'draft'
    data['is_active'] = False
    p = FeePlan(**data)
    await fee_plans_col.insert_one(p.model_dump())
    await log_audit(action='fee_plan.duplicate', current_user=current, school_id=plan['school_id'],
                    entity_type='fee_plan', entity_id=p.id,
                    details={'source': plan_id, 'name': p.name})
    return p.model_dump()


@api.delete('/fees/plans/{plan_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_fee_plan(plan_id: str, current=Depends(get_current_user)):
    plan = await fee_plans_col.find_one({'id': plan_id}, {'_id': 0})
    if not plan:
        raise HTTPException(404, 'Plan not found')
    used = await fee_assignments_col.count_documents({'fee_plan_id': plan_id})
    if used:
        raise HTTPException(400, f'Cannot delete: fee plan is used in {used} student assignment(s). Reassign those students first.')
    await fee_plans_col.delete_one({'id': plan_id})
    await log_audit(action='fee_plan.delete', current_user=current,
                    entity_type='fee_plan', entity_id=plan_id, details={'name': plan.get('name')})
    return {'ok': True}


@api.get('/fees/plans/{plan_id}/installments')
async def plan_installments(plan_id: str, current=Depends(get_current_user),
                            session: Optional[str] = None):
    """Return the plan's 12-month installment timeline (the MAIN fee structure's
    per-month amounts): 10-month split by default with June & March at ₹0,
    honouring per-month overrides, plan discounts and one-time charges.
    Used by the student Assign-Fee dialog to stay in sync with the structure."""
    plan = await fee_plans_col.find_one({'id': plan_id}, {'_id': 0})
    if not plan:
        raise HTTPException(404, 'Plan not found')
    if current['role'] != 'super_admin' and plan['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    sess = session or plan.get('academic_session') or '2026-27'
    return {'session': sess, 'installments': _build_plan_installments(plan, sess)}


@api.post('/fees/plans/{plan_id}/assign', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def assign_plan(plan_id: str, student_ids: List[str], request: Request,
                     current=Depends(get_current_user)):
    plan = await fee_plans_col.find_one({'id': plan_id}, {'_id': 0})
    if not plan:
        raise HTTPException(404, 'Plan not found')
    assignments = []
    for sid_ in student_ids:
        exists = await fee_assignments_col.find_one({'student_id': sid_, 'fee_plan_id': plan_id, 'academic_session': plan['academic_session']}, {'_id': 0})
        if exists:
            continue
        a = FeeAssignment(school_id=plan['school_id'], student_id=sid_, fee_plan_id=plan_id,
                          academic_session=plan['academic_session'])
        await fee_assignments_col.insert_one(a.model_dump())
        assignments.append(a.model_dump())
    await log_audit(action='fee_plan.assign', current_user=current, school_id=plan['school_id'],
                    entity_type='fee_plan', entity_id=plan_id,
                    details={'count': len(assignments)})
    return {'assigned': len(assignments)}


@api.get('/fees/student/{student_id}/dues')
async def student_dues(student_id: str, request: Request, current=Depends(get_current_user)):
    student = await students_col.find_one({'id': student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')
    assignments = await fee_assignments_col.find({'student_id': student_id}, {'_id': 0}).to_list(50)
    dues = []
    total_expected = 0.0
    total_discount = 0.0
    for a in assignments:
        # Custom items take priority over plan
        items_for_a = []
        if a.get('custom_items'):
            items_for_a = a['custom_items']
        elif a.get('fee_plan_id'):
            plan = await fee_plans_col.find_one({'id': a['fee_plan_id']}, {'_id': 0})
            if plan:
                items_for_a = plan.get('items', [])
        for item in items_for_a:
            entry = {
                'fee_head_id': item.get('fee_head_id'),
                'fee_head_name': item.get('fee_head_name'),
                'amount': item.get('amount', 0),
                'frequency': item.get('frequency', 'monthly'),
                'installments': item.get('installments', 1) if 'installments' in item else 1,
                'due_date': item.get('due_date') or a.get('due_date'),
                'assignment_id': a['id'],
                'plan_id': a.get('fee_plan_id'),
            }
            dues.append(entry)
            total_expected += entry['amount']
        # Apply assignment-level discount
        if a.get('discount_percent'):
            total_discount += total_expected * (a['discount_percent'] / 100)
        if a.get('discount_amount'):
            total_discount += a['discount_amount']
    # Plan-baked discounts (part of the plan; NO owner approval)
    for a in assignments:
        if not a.get('fee_plan_id') or a.get('custom_items'):
            continue
        plan = await fee_plans_col.find_one({'id': a['fee_plan_id']}, {'_id': 0})
        if not plan:
            continue
        pg = sum(float(it.get('amount') or 0) for it in (plan.get('items') or []))
        insts = a.get('installments') or []
        ams = [int(i['month']) for i in insts
               if (i.get('status') or 'active') != 'skip' and float(i.get('amount') or 0) > 0]
        if not ams:
            ams = [m['month'] for m in _session_months(a.get('academic_session') or '2026-27')
                   if m['month'] not in FEE_EXCLUDED_MONTHS]
        bd = compute_plan_discount_breakdown(plan, pg, ams)
        total_discount += bd['total_discount']
    # payments
    payments = await payments_col.find({'student_id': student_id, 'status': 'success'}, {'_id': 0}).to_list(500)
    total_paid = sum(p.get('total_paid', 0) for p in payments)
    # Approved per-payment discounts (owner-approved concessions granted at
    # collection time) also settle part of the fee, so they reduce the balance
    # exactly like paid money does.
    total_payment_discount = sum(float(p.get('discount') or 0) for p in payments)
    balance = max(total_expected - total_discount - total_payment_discount - total_paid, 0)

    # ---- Amount Due Till Date (monthly-payer aware) ----
    # If a student's parents pay monthly, we should NOT flag unpaid future
    # months as "due". Only count the pro-rated liability that has *elapsed*
    # by today (including the current month). This gives an accurate
    # "how much should have been paid by now" figure.
    net_annual = max(total_expected - total_discount, 0.0)
    # Pick the academic session — prefer the most-common one across assignments.
    session_counts_dt: Dict[str, int] = {}
    for a in assignments:
        s_ = a.get('academic_session')
        if s_:
            session_counts_dt[s_] = session_counts_dt.get(s_, 0) + 1
    current_session_dt = (max(session_counts_dt, key=session_counts_dt.get)
                          if session_counts_dt else '2026-27')
    months_list = _session_months(current_session_dt)
    today_dt_now = datetime.now()
    dues_installments = _merge_assignment_installments(assignments)
    if dues_installments:
        active_rows = [i for i in dues_installments
                       if (i.get('status') or 'active') != 'skip' and float(i.get('amount') or 0) > 0]
        months_total = len(active_rows)
        elapsed_rows = [i for i in active_rows
                        if (int(i.get('year') or 0), int(i.get('month') or 0)) <= (today_dt_now.year, today_dt_now.month)]
        months_elapsed = len(elapsed_rows)
        expected_till_date = round(sum(float(i.get('amount') or 0) for i in elapsed_rows), 2)
        # typical monthly amount = most common installment amount
        freq_map: Dict[float, int] = {}
        for i in active_rows:
            v = round(float(i.get('amount') or 0), 2)
            freq_map[v] = freq_map.get(v, 0) + 1
        monthly_amount = max(freq_map.items(), key=lambda kv: kv[1])[0] if freq_map else 0.0
    else:
        # Fees are split across 10 collection months (June & March excluded).
        collection_months_list = [m for m in months_list if m['month'] not in FEE_EXCLUDED_MONTHS]
        months_total = len(collection_months_list)
        months_elapsed = 0
        for m in collection_months_list:
            # include the current month once we are inside it
            if (m['year'], m['month']) <= (today_dt_now.year, today_dt_now.month):
                months_elapsed += 1
        months_elapsed = min(max(months_elapsed, 0), months_total)
        monthly_amount = round(net_annual / max(months_total, 1), 2) if net_annual > 0 else 0.0
        expected_till_date = round(monthly_amount * months_elapsed, 2)
    due_till_date = round(max(expected_till_date - total_paid - total_payment_discount, 0.0), 2)

    return {
        'student': student,
        'dues': dues,
        'assignments': assignments,
        'total_expected': total_expected,
        'total_discount': round(total_discount + total_payment_discount, 2),
        'total_plan_discount': round(total_discount, 2),
        'total_payment_discount': round(total_payment_discount, 2),
        'total_paid': total_paid,
        'balance': balance,
        'academic_session': current_session_dt,
        'monthly_amount': monthly_amount,
        'months_elapsed': months_elapsed,
        'months_total': months_total,
        'expected_till_date': expected_till_date,
        'due_till_date': due_till_date,
        'recent_payments': sorted(payments, key=lambda x: x.get('paid_at', ''), reverse=True)[:10],
    }


# ---------- MONTHLY / ANNUAL FEE SCHEDULE (for Parent Portal) ----------
def _session_months(session: str) -> list:
    """Return 12 (month_index, year, label) tuples for an academic session like '2026-27',
    starting from April of the first year."""
    try:
        start_year = int(session.split('-')[0])
    except Exception:
        start_year = datetime.now().year
    out = []
    for i in range(12):
        m = 4 + i  # April=4
        m2 = m if m <= 12 else m - 12
        y2 = start_year if m <= 12 else start_year + 1
        label = datetime(y2, m2, 1).strftime('%B %Y')
        out.append({'index': i, 'month': m2, 'year': y2, 'label': label})
    return out


def _merge_assignment_installments(assignments: list) -> Optional[list]:
    """Merge installments across ALL non-draft assignments into a single
    per-month timeline, SUMMING the amounts for each (year, month). This lets a
    student's monthly fee view show the TOTAL of every fee assigned in a month
    (e.g. recurring tuition + a one-time Books fee) instead of only one
    assignment. A month is 'skip'/No-Fee only when no assignment charges it.
    Returns None when no assignment has installments."""
    merged: Dict[tuple, float] = {}
    labels: Dict[tuple, str] = {}
    any_inst = False
    for a in assignments or []:
        if a.get('is_draft'):
            continue
        for i in (a.get('installments') or []):
            try:
                y, m = int(i['year']), int(i['month'])
            except Exception:
                continue
            any_inst = True
            status = i.get('status') or 'active'
            amt = float(i.get('amount') or 0)
            if status == 'skip' or amt <= 0:
                labels.setdefault((y, m), i.get('label') or 'No Fee')
                merged.setdefault((y, m), 0.0)
                continue
            merged[(y, m)] = round(merged.get((y, m), 0.0) + amt, 2)
    if not any_inst:
        return None
    out = []
    for (y, m), amt in merged.items():
        row = {'year': y, 'month': m, 'amount': round(amt, 2),
               'status': 'active' if amt > 0 else 'skip'}
        if amt <= 0 and (y, m) in labels:
            row['label'] = labels[(y, m)]
        out.append(row)
    return out


def _pick_assignment_installments(assignments: list) -> Optional[list]:
    """Return the installments array of the most recently created NON-DRAFT
    assignment that has one. Falls back to any assignment with installments.
    Returns None when nothing usable exists."""
    best = None
    best_key = ('', 0)  # (created_at, non_draft_priority)
    for a in assignments or []:
        inst = a.get('installments') or []
        if not inst:
            continue
        key = (a.get('created_at') or '', 0 if a.get('is_draft') else 1)
        # prefer non-draft; among same draft-ness, prefer latest created_at
        if best is None or (key[1], key[0]) > (best_key[1], best_key[0]):
            best = inst
            best_key = key
    return best


# ---------- PLAN-BAKED DISCOUNT COMPUTATION (no owner approval) ----------
def _discount_amount(dtype: Optional[str], value, base: float) -> float:
    """Resolve a single discount (flat ₹ or percent %) against `base`.
    Returns a non-negative amount, capped at `base`. A percent is clamped 0..100."""
    if not dtype:
        return 0.0
    try:
        value = float(value or 0)
    except Exception:
        return 0.0
    if value <= 0 or base <= 0:
        return 0.0
    if dtype == 'percent':
        pct = min(max(value, 0.0), 100.0)
        return round(base * pct / 100.0, 2)
    return round(min(value, base), 2)  # flat


def _plan_month_discount_map(plan: dict) -> Dict[int, dict]:
    """Merge a plan's month_discounts + installment_discounts into one
    {month: {'type','value'}} map. `month_discounts` win when a month appears
    in both — this avoids applying two per-month discounts to the same month."""
    out: Dict[int, dict] = {}
    for md in (plan.get('installment_discounts') or []):
        try:
            out[int(md.get('month'))] = {'type': md.get('type', 'flat'), 'value': float(md.get('value') or 0)}
        except Exception:
            continue
    for md in (plan.get('month_discounts') or []):
        try:
            out[int(md.get('month'))] = {'type': md.get('type', 'flat'), 'value': float(md.get('value') or 0)}
        except Exception:
            continue
    return out


def compute_plan_discount_breakdown(plan: dict, annual_gross: float,
                                    active_months: Optional[list] = None) -> dict:
    """Central plan-discount computation. Discount resolution priority
    (documented, no double-applying):
      1. Plan-level discount (flat/percent) reduces the ANNUAL gross.
      2. Yearly/full-session discount (flat/percent) reduces the remaining annual amount.
         -> net-after-lump, split equally across active months = base month amount.
      3. Month-specific discounts (flat/percent) reduce ONLY that month's base amount.
         They apply at the month level (stack on top of the annual-level cuts), so no
         single month is discounted twice by the same rule.
    Returns gross, each discount bucket, net_annual and per-month rows.
    """
    annual_gross = round(float(annual_gross or 0), 2)
    plan_disc = _discount_amount(plan.get('plan_discount_type'), plan.get('plan_discount_value'), annual_gross)
    remaining = max(annual_gross - plan_disc, 0.0)
    yearly_disc = _discount_amount(plan.get('yearly_discount_type'), plan.get('yearly_discount_value'), remaining)
    net_after_lump = max(annual_gross - plan_disc - yearly_disc, 0.0)

    months = list(active_months) if active_months else list(range(1, 13))
    n = len(months) or 1
    base_month = round(net_after_lump / n, 2) if net_after_lump > 0 else 0.0
    md_map = _plan_month_discount_map(plan)

    rows = []
    month_disc_total = 0.0
    for m in months:
        md = md_map.get(int(m))
        d = _discount_amount(md.get('type'), md.get('value'), base_month) if md else 0.0
        month_disc_total += d
        rows.append({'month': int(m), 'base': base_month, 'discount': round(d, 2),
                     'net': max(round(base_month - d, 2), 0.0)})
    month_disc_total = round(month_disc_total, 2)
    total_discount = round(plan_disc + yearly_disc + month_disc_total, 2)
    net_annual = max(round(annual_gross - total_discount, 2), 0.0)
    return {
        'annual_gross': annual_gross,
        'plan_discount_amount': plan_disc,
        'yearly_discount_amount': yearly_disc,
        'month_discount_total': month_disc_total,
        'total_discount': total_discount,
        'net_annual': net_annual,
        'base_month': base_month,
        'per_month': rows,
    }


def _validate_plan_discounts(data: dict) -> None:
    """Validate plan discount fields: type in {flat,percent}, no negatives,
    percent 0..100, month 1..12. Raises HTTPException(400) on any violation."""
    def chk(dtype, value, label):
        if dtype is None:
            return
        if dtype not in ('flat', 'percent'):
            raise HTTPException(400, f'{label}: type must be "flat" or "percent"')
        try:
            v = float(value or 0)
        except Exception:
            raise HTTPException(400, f'{label}: value must be a number')
        if v < 0:
            raise HTTPException(400, f'{label}: value cannot be negative')
        if dtype == 'percent' and v > 100:
            raise HTTPException(400, f'{label}: percent must be between 0 and 100')

    chk(data.get('plan_discount_type'), data.get('plan_discount_value'), 'Plan discount')
    chk(data.get('yearly_discount_type'), data.get('yearly_discount_value'), 'Yearly discount')
    for key in ('month_discounts', 'installment_discounts'):
        for md in (data.get(key) or []):
            try:
                m = int(md.get('month') or 0)
            except Exception:
                raise HTTPException(400, 'Month discount: month must be an integer 1..12')
            if m < 1 or m > 12:
                raise HTTPException(400, 'Month discount: month must be between 1 and 12')
            chk(md.get('type'), md.get('value'), f'Month {m} discount')
    for ma in (data.get('month_amounts') or []):
        try:
            mm = int(ma.get('month') or 0)
        except Exception:
            raise HTTPException(400, 'Month amount: month must be an integer 1..12')
        if mm < 1 or mm > 12:
            raise HTTPException(400, 'Month amount: month must be between 1 and 12')
        try:
            av = float(ma.get('amount') or 0)
        except Exception:
            raise HTTPException(400, 'Month amount: amount must be a number')
        if av < 0:
            raise HTTPException(400, 'Month amount: amount cannot be negative')



def _build_month_schedule(net_annual: float, payments: list, session: str,
                          today: Optional[datetime] = None, grace_day: int = 15,
                          installments: Optional[list] = None,
                          extra_paid_credit: float = 0.0) -> tuple:
    """Compute the 12-month schedule for a student given their net annual fee
    liability and the list of *success* payments. Returns
    (schedule, monthly_amount, fully_paid, overdue_count, overdue_amount, due_amount).

    - When the fee assignment carries an explicit `installments` array, the
      per-month amounts come from it (e.g. 10-month split with June & March at
      ₹0, or a one-time charge in a single month). Months with zero amount are
      marked `no_fee`.
    - Otherwise the net annual amount is split equally across 12 months.
    - Explicit line-item `period` labels (e.g. "April 2026") mark those months as paid.
    - Remaining paid amount is absorbed FIFO across the still-pending months.
    - Months whose (year, month) is strictly before the current month AND still
      unpaid/partial are flagged `overdue`. Additionally, the CURRENT month is
      also treated as overdue once today's day exceeds `grace_day` (default 15th).
    """
    total_paid = sum(float(p.get('total_paid') or 0) for p in payments) + float(extra_paid_credit or 0)

    # ---- per-month amount map from installments (if provided) ----
    inst_map: Dict[tuple, dict] = {}
    for i in (installments or []):
        try:
            inst_map[(int(i.get('year')), int(i.get('month')))] = i
        except Exception:
            continue

    explicit_paid_labels = set()
    for p in payments:
        for it in (p.get('items') or []):
            per = (it.get('period') or '').strip()
            if per:
                explicit_paid_labels.add(per)

    # Fees split across 10 collection months only (June & March excluded).
    equal_monthly = round(net_annual / max(FEE_COLLECTION_MONTH_COUNT, 1), 2) if net_annual > 0 else 0.0
    months = _session_months(session)
    schedule = []
    explicit_absorbed = 0.0
    for m in months:
        if inst_map:
            inst = inst_map.get((m['year'], m['month']))
            amt = 0.0
            if inst and (inst.get('status') or 'active') != 'skip':
                amt = round(float(inst.get('amount') or 0), 2)
            status = 'pending' if amt > 0 else 'no_fee'
        elif m['month'] in FEE_EXCLUDED_MONTHS:
            amt = 0.0
            status = 'no_fee'
        else:
            amt = equal_monthly
            status = 'pending'
        entry = {
            'index': m['index'], 'label': m['label'],
            'month': m['month'], 'year': m['year'],
            'amount': amt, 'paid_amount': 0.0, 'status': status,
        }
        if status != 'no_fee' and m['label'] in explicit_paid_labels:
            entry['status'] = 'paid'
            entry['paid_amount'] = amt
            explicit_absorbed += amt
        schedule.append(entry)

    remaining = max(total_paid - explicit_absorbed, 0.0)
    for entry in schedule:
        if entry['status'] in ('paid', 'no_fee'):
            continue
        if remaining <= 0:
            break
        if remaining >= entry['amount'] - 0.01:
            entry['status'] = 'paid'
            entry['paid_amount'] = entry['amount']
            remaining -= entry['amount']
        else:
            entry['status'] = 'partial'
            entry['paid_amount'] = round(remaining, 2)
            remaining = 0

    today = today or datetime.now()
    overdue_count = 0
    overdue_amount = 0.0
    for entry in schedule:
        if entry['status'] in ('pending', 'partial'):
            # 1) Any month strictly before today's (year, month) — clearly overdue.
            # 2) The current month becomes overdue once today's day > grace_day.
            is_past_month = (entry['year'], entry['month']) < (today.year, today.month)
            is_current_month_past_grace = (
                entry['year'] == today.year and entry['month'] == today.month
                and today.day > grace_day
            )
            if is_past_month or is_current_month_past_grace:
                if entry['status'] == 'pending':
                    entry['status'] = 'overdue'
                overdue_count += 1
                overdue_amount += max(entry['amount'] - entry['paid_amount'], 0.0)

    # "Typical" monthly amount — the most common non-zero installment amount,
    # falling back to equal 12-way split.
    active_amounts = [e['amount'] for e in schedule if e['status'] != 'no_fee' and e['amount'] > 0]
    if active_amounts:
        freq: Dict[float, int] = {}
        for v in active_amounts:
            freq[v] = freq.get(v, 0) + 1
        monthly_amount = max(freq.items(), key=lambda kv: kv[1])[0]
    else:
        monthly_amount = equal_monthly

    fully_paid = net_annual > 0 and all(e['status'] in ('paid', 'no_fee') for e in schedule)
    due_amount = round(max(net_annual - total_paid, 0.0), 2)
    return schedule, monthly_amount, fully_paid, overdue_count, round(overdue_amount, 2), due_amount


@api.get('/fees/student/{student_id}/fee-schedule')
async def student_fee_schedule(student_id: str, request: Request, current=Depends(get_current_user)):
    """Return a monthly & annual view of the student's fee liability for the
    parent portal. Splits the annual tuition/etc. into 12 equal monthly parts,
    marks months as paid based on prior payments (FIFO), and computes the
    'pay full' amount with the plan's annual discount applied."""
    student = await students_col.find_one({'id': student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')

    assignments = await fee_assignments_col.find({'student_id': student_id}, {'_id': 0}).to_list(50)

    # Aggregate: annual gross, per-assignment concession, session, and the
    # primary fee plan (used for plan-baked discounts) from the assignments.
    annual_total = 0.0
    concession = 0.0            # per-assignment ad-hoc discount (owner-approval path)
    annual_discount_percent = 0.0
    session = '2026-27'
    items_flat: list = []       # for reference display
    primary_plan = None
    for a in assignments:
        session = a.get('academic_session') or session
        concession += float(a.get('discount_amount') or 0)
        if a.get('custom_items'):
            for it in a['custom_items']:
                items_flat.append(it)
                annual_total += float(it.get('amount') or 0)
        elif a.get('fee_plan_id'):
            plan = await fee_plans_col.find_one({'id': a['fee_plan_id']}, {'_id': 0})
            if plan:
                primary_plan = primary_plan or plan
                annual_discount_percent = max(annual_discount_percent, float(plan.get('annual_discount_percent') or 0))
                for it in (plan.get('items') or []):
                    items_flat.append(it)
                    annual_total += float(it.get('amount') or 0)
        if a.get('discount_percent'):
            concession += (annual_total * float(a['discount_percent']) / 100.0)

    annual_total = round(annual_total, 2)
    concession = round(concession, 2)

    # ---- Plan-baked discounts (part of the plan; NO owner approval) ----
    assignment_installments = _merge_assignment_installments(assignments)
    if assignment_installments:
        active_month_nums = [int(i['month']) for i in assignment_installments
                             if (i.get('status') or 'active') != 'skip' and float(i.get('amount') or 0) > 0]
    else:
        active_month_nums = [m['month'] for m in _session_months(session)
                             if m['month'] not in FEE_EXCLUDED_MONTHS]

    plan_lump = 0.0
    month_disc_map: Dict[int, dict] = {}
    if primary_plan:
        pb = compute_plan_discount_breakdown(primary_plan, annual_total, active_month_nums)
        plan_lump = round(pb['plan_discount_amount'] + pb['yearly_discount_amount'], 2)
        month_disc_map = _plan_month_discount_map(primary_plan)

    net_after_lump = max(round(annual_total - concession - plan_lump, 2), 0.0)

    # Build per-month base installments (explicit or equal split), then apply
    # plan month-specific discounts to individual months.
    months = _session_months(session)
    if assignment_installments:
        base_installments = [dict(i) for i in assignment_installments]
    else:
        per = round(net_after_lump / max(len(active_month_nums), 1), 2) if net_after_lump > 0 else 0.0
        base_installments = [
            {'month': m['month'], 'year': m['year'],
             'amount': (0.0 if m['month'] in FEE_EXCLUDED_MONTHS else per),
             'status': ('skip' if m['month'] in FEE_EXCLUDED_MONTHS else 'active')}
            for m in months
        ]

    month_discount_total = 0.0
    if assignment_installments:
        # Explicit installments already encode ALL plan discounts (plan, yearly
        # and per-month) — do NOT re-apply them. Net annual = sum of active
        # installment amounts minus the per-student concession. The month-
        # discount total is read from the plan breakdown purely for display.
        inst_sum = sum(float(i.get('amount') or 0) for i in base_installments
                       if (i.get('status') or 'active') != 'skip')
        if primary_plan:
            month_discount_total = round(compute_plan_discount_breakdown(
                primary_plan, annual_total, active_month_nums)['month_discount_total'], 2)
        net_annual = max(round(inst_sum - concession, 2), 0.0)
    else:
        if month_disc_map:
            for inst in base_installments:
                if (inst.get('status') or 'active') == 'skip':
                    continue
                md = month_disc_map.get(int(inst['month']))
                if not md:
                    continue
                d = _discount_amount(md.get('type'), md.get('value'), float(inst.get('amount') or 0))
                if d > 0:
                    inst['amount'] = round(max(float(inst.get('amount') or 0) - d, 0.0), 2)
                    month_discount_total += d
        month_discount_total = round(month_discount_total, 2)
        net_annual = max(round(net_after_lump - month_discount_total, 2), 0.0)

    # Payments so far
    payments = await payments_col.find({'student_id': student_id, 'status': 'success'}, {'_id': 0}).to_list(500)
    total_paid = sum(float(p.get('total_paid') or 0) for p in payments)
    # Owner-approved per-payment discounts settle part of the fee too — credit
    # them against the balance and the month schedule just like paid money.
    total_payment_discount = sum(float(p.get('discount') or 0) for p in payments)

    # Build 12-month schedule (shared helper) — honours the assignment's
    # explicit installments and the plan's discounted per-month amounts.
    schedule, monthly_amount, _fully_paid, _oc, _oa, _due = _build_month_schedule(
        net_annual, payments, session, installments=base_installments,
        extra_paid_credit=total_payment_discount,
    )
    active_months = sum(1 for e in schedule if e['status'] != 'no_fee')

    remaining_balance = max(net_annual - total_paid - total_payment_discount, 0.0)
    # Pay Full discount — apply annual_discount_percent on the REMAINING amount.
    full_payment_discount = round(remaining_balance * (annual_discount_percent / 100.0), 2) if annual_discount_percent > 0 else 0.0
    payable_full = round(max(remaining_balance - full_payment_discount, 0.0), 2)

    plan_discount_total = round(plan_lump + month_discount_total, 2)
    you_saved = round(concession + plan_discount_total + total_payment_discount, 2)

    return {
        'student': student,
        'academic_session': session,
        'annual_total': annual_total,
        'concession': concession,
        'plan_discount_total': plan_discount_total,
        'plan_lump_discount': plan_lump,
        'month_discount_total': month_discount_total,
        'you_saved': you_saved,
        'net_annual': round(net_annual, 2),
        'monthly_amount': monthly_amount,
        'active_months': active_months,
        'total_paid': round(total_paid, 2),
        'payment_discount_total': round(total_payment_discount, 2),
        'remaining_balance': round(remaining_balance, 2),
        'annual_discount_percent': annual_discount_percent,
        'full_payment_discount': full_payment_discount,
        'payable_full': payable_full,
        'schedule': schedule,
        'fee_head_names': list({(it.get('fee_head_name') or 'Tuition Fee') for it in items_flat}) or ['Tuition Fee'],
    }


# ---------- FEE ASSIGNMENTS (per-student) ----------
@api.get('/fees/assignments')
async def list_assignments(request: Request, current=Depends(get_current_user),
                           school_id: Optional[str] = None,
                           student_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if student_id:
        q['student_id'] = student_id
    return await fee_assignments_col.find(q, {'_id': 0}).to_list(500)


@api.get('/fees/assignments/previous-for-student/{student_id}')
async def previous_assignment_for_student(student_id: str, request: Request,
                                          current=Depends(get_current_user)):
    """Return the most recent non-draft fee assignment for a student in a
    PRIOR academic session — used by the AssignFees dialog's
    "Copy Previous Year's Fee Structure" button.
    Response is either the assignment doc (with a `source_session` echo) or
    { previous: null } when there's nothing to copy.
    """
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')
    # Determine the "current" session: highest session across assignments for
    # this student. Then look for anything with a different (older) session.
    docs = await fee_assignments_col.find(
        {'student_id': student_id, 'is_draft': {'$ne': True}},
        {'_id': 0},
    ).sort('created_at', -1).to_list(50)
    if not docs:
        return {'previous': None}
    latest_session = docs[0].get('academic_session')
    prev = next((d for d in docs if d.get('academic_session') and d.get('academic_session') != latest_session), None)
    if not prev:
        return {'previous': None}
    return {'previous': prev, 'source_session': prev.get('academic_session')}


@api.post('/fees/assignments', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_assignment(body: FeeAssignmentCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    installments_input = body.installments or []
    a = FeeAssignment(
        school_id=sid, student_id=body.student_id, fee_plan_id=body.fee_plan_id,
        academic_session=body.academic_session,
        custom_items=body.custom_items or [],
        custom_amount=body.custom_amount,
        discount_percent=body.discount_percent or 0.0,
        discount_amount=body.discount_amount or 0.0,
        discount_reason=body.discount_reason,
        due_date=body.due_date, remarks=body.remarks,
        internal_notes=body.internal_notes,
        collection_months=body.collection_months or [4, 5, 7, 8, 9, 10, 11, 12, 1, 2],
        installments=[FeeAssignmentInstallment(**i.model_dump()) if hasattr(i, 'model_dump') else FeeAssignmentInstallment(**i) for i in installments_input],
        due_day_of_month=body.due_day_of_month or 15,
        is_draft=bool(body.is_draft),
        copied_from_assignment_id=body.copied_from_assignment_id,
    )
    await fee_assignments_col.insert_one(a.model_dump())
    await log_audit(action='fee_assignment.create', current_user=current, school_id=sid,
                    entity_type='fee_assignment', entity_id=a.id,
                    details={'student_id': body.student_id, 'discount_percent': body.discount_percent,
                             'discount_amount': body.discount_amount, 'due_date': body.due_date,
                             'is_draft': a.is_draft})
    # Optionally push a notification to linked parents so the mobile portal
    # gets a bell/badge alert. Only fires for non-draft, notify_parent==True.
    if body.notify_parent and not a.is_draft:
        try:
            await _create_fee_assignment_notification(sid, body.student_id, a, current)
        except Exception as e:
            # Never fail the assignment because of a notification hiccup.
            logging.exception('fee-assignment notification failed: %s', e)
    return a.model_dump()


async def _create_fee_assignment_notification(sid: str, student_id: str, a: FeeAssignment, current: Dict[str, Any]) -> None:
    student = await students_col.find_one({'id': student_id}, {'_id': 0}) or {}
    plan = None
    if a.fee_plan_id:
        plan = await fee_plans_col.find_one({'id': a.fee_plan_id}, {'_id': 0})
    # Compute gross/net for the message body.
    if plan:
        gross = sum(float(it.get('amount') or 0) for it in (plan.get('items') or []))
    else:
        gross = sum(float(it.amount or 0) for it in a.custom_items)
    if a.discount_amount and a.discount_amount > 0:
        disc = float(a.discount_amount)
    else:
        disc = round(gross * float(a.discount_percent or 0) / 100, 2)
    net = max(gross - disc, 0)
    active_months = [i for i in a.installments if i.status == 'active'] if a.installments else []
    monthly = round(net / max(len(active_months) or (len(a.collection_months) or 12), 1), 2)
    title = f"New fee assigned — {student.get('full_name') or 'your child'}"
    body_text = (
        f"An annual fee plan of ₹{gross:,.2f}"
        + (f" (discount ₹{disc:,.2f})" if disc > 0 else '')
        + f" — Net Payable ₹{net:,.2f} — has been assigned"
        + (f" ({len(active_months)} monthly instalments of ₹{monthly:,.2f})" if active_months else '')
        + ". Please log in to the Parent Portal for the month-wise schedule."
    )
    n = Notification(
        school_id=sid, title=title, body=body_text,
        audience='parents', class_id=student.get('class_id'),
        student_ids=[student_id], kind='fee_reminder',
    )
    await notifications_col.insert_one(n.model_dump())
    await log_audit(action='notification.create', current_user=current, school_id=sid,
                    entity_type='notification', entity_id=n.id,
                    details={'title': n.title, 'audience': n.audience,
                             'trigger': 'fee_assignment', 'assignment_id': a.id})


@api.patch('/fees/assignments/{assignment_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_assignment(assignment_id: str, body: FeeAssignmentUpdate, current=Depends(get_current_user)):
    upd_raw = body.model_dump(exclude_none=True)
    # notify_parent is transient — never persisted on the assignment.
    notify_parent = bool(upd_raw.pop('notify_parent', False))
    upd_raw['updated_at'] = now_iso()
    r = await fee_assignments_col.update_one({'id': assignment_id}, {'$set': upd_raw})
    if not r.matched_count:
        raise HTTPException(404, 'Assignment not found')
    await log_audit(action='fee_assignment.update', current_user=current,
                    entity_type='fee_assignment', entity_id=assignment_id, details=upd_raw)
    doc = await fee_assignments_col.find_one({'id': assignment_id}, {'_id': 0})
    if notify_parent and doc and not doc.get('is_draft'):
        try:
            a = FeeAssignment(**doc)
            await _create_fee_assignment_notification(doc['school_id'], doc['student_id'], a, current)
        except Exception as e:
            logging.exception('fee-assignment notification (patch) failed: %s', e)
    return doc


@api.delete('/fees/assignments/{assignment_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_assignment(assignment_id: str, current=Depends(get_current_user)):
    r = await fee_assignments_col.delete_one({'id': assignment_id})
    if not r.deleted_count:
        raise HTTPException(404, 'Assignment not found')
    await log_audit(action='fee_assignment.delete', current_user=current,
                    entity_type='fee_assignment', entity_id=assignment_id)
    return {'ok': True}


# =====================================================
# PAYMENTS / RECEIPTS
# =====================================================
async def _generate_receipt_number(school_id: str) -> str:
    school = await schools_col.find_one({'id': school_id}, {'_id': 0})
    code = (school or {}).get('code', 'STV')
    seq = await get_next_sequence(f'receipt_{school_id}')
    year = datetime.now().year
    return f'{code}-{year}-{seq:06d}'


async def _finalize_payment(school_id: str, body: PaymentCreate, current: dict,
                            razorpay_order_id: Optional[str] = None,
                            razorpay_payment_id: Optional[str] = None,
                            razorpay_signature: Optional[str] = None) -> Payment:
    student = await students_col.find_one({'id': body.student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    subtotal = sum(i.amount for i in body.items)
    total_paid = subtotal + body.late_fee - body.discount
    receipt_no = await _generate_receipt_number(school_id)
    payment = Payment(
        school_id=school_id,
        student_id=student['id'],
        student_name=student['full_name'],
        receipt_number=receipt_no,
        items=body.items,
        subtotal=subtotal,
        discount=body.discount,
        late_fee=body.late_fee,
        total_paid=total_paid,
        payment_mode=body.payment_mode,
        txn_ref=body.txn_ref,
        razorpay_order_id=razorpay_order_id,
        razorpay_payment_id=razorpay_payment_id,
        razorpay_signature=razorpay_signature,
        status='success',
        remarks=body.remarks,
        collected_by_id=current.get('id'),
        collected_by_name=current.get('full_name'),
    )
    await payments_col.insert_one(payment.model_dump())
    await log_audit(action='payment.collect', current_user=current, school_id=school_id,
                    entity_type='payment', entity_id=payment.id,
                    details={'receipt_number': receipt_no, 'total': total_paid,
                             'mode': body.payment_mode, 'student': student['full_name']})
    return payment


@api.post('/payments/collect', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant'))])
async def collect_payment(body: PaymentCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    # Accountants are NOT permitted to apply discounts. They can only collect
    # the exact fee due. Any discount must be initiated by a School Admin or
    # Super Admin, and still requires Owner approval before receipt is generated.
    if current.get('role') == 'accountant' and float(body.discount or 0) > 0:
        raise HTTPException(
            403,
            'Accountants are not permitted to apply discounts. Please ask a School Admin or Super Admin to raise a discount request.'
        )
    # Discounted payments ALWAYS require owner approval before receipt is
    # generated. There is no bypass for super_admin / school_admin — the whole
    # point of the workflow is that the school owner sees every discount.
    # Owners themselves do not collect fees (UI does not expose the endpoint
    # to them); if they ever hit it directly, they too must submit for peer
    # approval — which is fine because owners are on both sides of the ledger.
    if float(body.discount or 0) > 0:
        if not (body.discount_reason and body.discount_reason.strip()):
            raise HTTPException(400, 'Discount reason is required when a discount is applied')
        # Application image is MANDATORY when a discount is requested — the
        # parent's written application on plain paper must be scanned/photographed
        # and attached so the owner can review the justification.
        img = (body.application_image or '').strip()
        if not img:
            raise HTTPException(
                400,
                "An image of the parent's written application is required to raise a discount approval request."
            )
        # Basic sanity: must look like a base64 data URL for an image or PDF.
        if not (img.startswith('data:image/') or img.startswith('data:application/pdf')):
            raise HTTPException(400, 'Application attachment must be an image or PDF (base64 data URL).')
        # Size guard — 3 MB decoded max (approx 4 MB base64) to keep DB docs sane.
        try:
            _, b64 = img.split(',', 1)
            approx_bytes = int(len(b64) * 3 / 4)
        except Exception:
            approx_bytes = len(img)
        if approx_bytes > 3 * 1024 * 1024:
            raise HTTPException(400, 'Application attachment is too large (max 3 MB).')
        student = await students_col.find_one({'id': body.student_id}, {'_id': 0})
        if not student:
            raise HTTPException(404, 'Student not found')
        klass = None
        if student.get('class_id'):
            klass = await classes_col.find_one({'id': student['class_id']}, {'_id': 0})
        subtotal = sum(i.amount for i in body.items)
        total = subtotal + float(body.late_fee or 0) - float(body.discount or 0)
        approval = DiscountApproval(
            school_id=sid,
            student_id=student['id'],
            student_name=student.get('full_name', ''),
            class_id=student.get('class_id'),
            class_name=(klass or {}).get('name'),
            section=student.get('section'),
            admission_number=student.get('admission_number'),
            items=body.items,
            subtotal=subtotal,
            discount=float(body.discount or 0),
            late_fee=float(body.late_fee or 0),
            total=total,
            payment_mode=body.payment_mode,
            txn_ref=body.txn_ref,
            remarks=body.remarks,
            discount_reason=body.discount_reason.strip(),
            application_image=img,
            status='pending',
            requested_by_id=current.get('id'),
            requested_by_name=current.get('full_name'),
            requested_by_role=current.get('role'),
        )
        await discount_approvals_col.insert_one(approval.model_dump())
        await log_audit(action='discount_approval.request', current_user=current, school_id=sid,
                        entity_type='discount_approval', entity_id=approval.id,
                        details={'student': student.get('full_name'), 'discount': approval.discount,
                                 'reason': approval.discount_reason})
        return {
            'status': 'pending_approval',
            'approval_id': approval.id,
            'message': 'Discount requires owner approval. Once approved, collect the money and generate the receipt.',
            'approval': approval.model_dump(),
        }
    # No discount — collect immediately
    payment = await _finalize_payment(sid, body, current)
    return payment.model_dump()


# ================= DISCOUNT APPROVAL ENDPOINTS =================
@api.get('/discount-approvals')
async def list_discount_approvals(request: Request,
                                  status: Optional[str] = None,
                                  current=Depends(get_current_user)):
    role = current.get('role')
    if role not in ('owner', 'super_admin', 'school_admin', 'accountant'):
        raise HTTPException(403, 'Forbidden')
    query: Dict[str, Any] = {}
    if status:
        query['status'] = status
    # Scope by school for non super_admin
    if role != 'super_admin':
        sid = current.get('school_id')
        # Owners: if they have a school_id, scope; else show all (multi-branch owners).
        if role != 'owner' or sid:
            if sid:
                query['school_id'] = sid
        # school_admin/accountant only see their own submissions
        if role in ('school_admin', 'accountant'):
            query['requested_by_id'] = current.get('id')
    else:
        x_sid = request.headers.get('X-School-Id')
        if x_sid:
            query['school_id'] = x_sid
    docs = await discount_approvals_col.find(query, {'_id': 0}).sort('requested_at', -1).to_list(length=1000)
    return docs


@api.get('/discount-approvals/pending-count')
async def pending_discount_approvals_count(request: Request, current=Depends(get_current_user)):
    role = current.get('role')
    if role not in ('owner', 'super_admin', 'school_admin', 'accountant'):
        return {'count': 0}
    query: Dict[str, Any] = {'status': 'pending'}
    if role != 'super_admin':
        sid = current.get('school_id')
        if sid and role != 'owner':
            query['school_id'] = sid
        elif sid and role == 'owner':
            query['school_id'] = sid
        if role in ('school_admin', 'accountant'):
            query['requested_by_id'] = current.get('id')
    else:
        x_sid = request.headers.get('X-School-Id')
        if x_sid:
            query['school_id'] = x_sid
    count = await discount_approvals_col.count_documents(query)
    return {'count': count}


@api.get('/discount-approvals/awaiting-collection-count')
async def awaiting_collection_count(request: Request, current=Depends(get_current_user)):
    """Count of approved-but-not-yet-collected discount approvals — used to
    show the admin a small badge on Fee Collection so they know there are
    approvals waiting to be turned into receipts."""
    role = current.get('role')
    if role not in ('super_admin', 'school_admin', 'accountant', 'owner'):
        return {'count': 0}
    query: Dict[str, Any] = {'status': 'approved'}
    if role != 'super_admin':
        sid = current.get('school_id')
        if sid:
            query['school_id'] = sid
        if role in ('school_admin', 'accountant'):
            query['requested_by_id'] = current.get('id')
    else:
        x_sid = request.headers.get('X-School-Id')
        if x_sid:
            query['school_id'] = x_sid
    count = await discount_approvals_col.count_documents(query)
    return {'count': count}


@api.get('/discount-approvals/{approval_id}')
async def get_discount_approval(approval_id: str, current=Depends(get_current_user)):
    role = current.get('role')
    if role not in ('owner', 'super_admin', 'school_admin', 'accountant'):
        raise HTTPException(403, 'Forbidden')
    doc = await discount_approvals_col.find_one({'id': approval_id}, {'_id': 0})
    if not doc:
        raise HTTPException(404, 'Approval not found')
    # school_admin/accountant may only see own
    if role in ('school_admin', 'accountant') and doc.get('requested_by_id') != current.get('id'):
        raise HTTPException(403, 'Forbidden')
    return doc


@api.post('/discount-approvals/{approval_id}/approve',
          dependencies=[Depends(require_roles('owner'))])
async def approve_discount_request(approval_id: str, body: DiscountApprovalReview,
                                   current=Depends(get_current_user)):
    """STAGE 2 of the two-stage discount workflow.
    Owner reviews the request (with the parent's written application image)
    and approves it, optionally overriding the requested discount amount.
    NO Payment / receipt is generated here — the original admin must resume
    the transaction (POST /discount-approvals/{id}/collect) after collecting
    the money from the parent.
    """
    doc = await discount_approvals_col.find_one({'id': approval_id}, {'_id': 0})
    if not doc:
        raise HTTPException(404, 'Approval not found')
    if doc.get('status') != 'pending':
        raise HTTPException(400, f"Approval already {doc.get('status')}")

    # Owner may override the requested discount amount. If not supplied, use
    # the original requested value. Clamp within [0, subtotal + late_fee].
    subtotal = float(doc.get('subtotal') or 0)
    late_fee = float(doc.get('late_fee') or 0)
    requested_discount = float(doc.get('discount') or 0)
    if body.approved_discount is not None:
        try:
            approved_discount = float(body.approved_discount)
        except Exception:
            raise HTTPException(400, 'approved_discount must be a number')
        if approved_discount < 0:
            raise HTTPException(400, 'approved_discount cannot be negative')
        # A zero discount here would defeat the purpose — force > 0
        if approved_discount == 0:
            raise HTTPException(
                400,
                'To approve with zero discount, reject the request and ask the admin to collect at full amount.'
            )
        if approved_discount > (subtotal + late_fee):
            raise HTTPException(400, 'approved_discount cannot exceed subtotal + late fee')
    else:
        approved_discount = requested_discount

    approved_total = round(subtotal + late_fee - approved_discount, 2)

    await discount_approvals_col.update_one(
        {'id': approval_id},
        {'$set': {
            'status': 'approved',
            'approved_discount': approved_discount,
            'total': approved_total,
            'reviewed_by_id': current.get('id'),
            'reviewed_by_name': current.get('full_name'),
            'reviewed_at': now_iso(),
            'review_remark': (body.remark or '').strip() or None,
        }}
    )
    await log_audit(action='discount_approval.approve', current_user=current,
                    school_id=doc['school_id'], entity_type='discount_approval',
                    entity_id=approval_id,
                    details={
                        'student': doc.get('student_name'),
                        'requested_discount': requested_discount,
                        'approved_discount': approved_discount,
                    })
    updated = await discount_approvals_col.find_one({'id': approval_id}, {'_id': 0})
    return {
        'ok': True,
        'status': 'approved',
        'approved_discount': approved_discount,
        'total': approved_total,
        'approval': updated,
        'message': 'Approved. The admin can now collect the money and generate the receipt.',
    }


@api.post('/discount-approvals/{approval_id}/collect',
          dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant'))])
async def collect_discount_approval(approval_id: str, body: DiscountApprovalCollect,
                                    request: Request, current=Depends(get_current_user)):
    """STAGE 3 of the two-stage discount workflow.
    After the owner has APPROVED the discount, the admin returns here to
    actually collect the money and generate the fee receipt. Only then is
    a Payment document created.
    """
    doc = await discount_approvals_col.find_one({'id': approval_id}, {'_id': 0})
    if not doc:
        raise HTTPException(404, 'Approval not found')
    if doc.get('status') != 'approved':
        raise HTTPException(400, f"Cannot collect — approval status is '{doc.get('status')}'.")
    if doc.get('payment_id'):
        raise HTTPException(400, 'This approval has already been collected.')

    # School scoping: only staff of the same school can collect (super_admin
    # is unrestricted). This keeps multi-branch cleanliness.
    if current.get('role') != 'super_admin':
        user_sid = current.get('school_id')
        if user_sid and doc.get('school_id') != user_sid:
            raise HTTPException(403, 'You cannot collect approvals for another school.')

    # Use the owner-approved discount (fallback to originally requested).
    approved_discount = doc.get('approved_discount')
    if approved_discount is None:
        approved_discount = doc.get('discount', 0.0)
    approved_discount = float(approved_discount or 0)

    # Build a PaymentCreate from the approval, using overridden payment_mode /
    # txn_ref / remarks from the request body if provided.
    payment_mode = (body.payment_mode or doc.get('payment_mode') or 'cash').strip()
    txn_ref = body.txn_ref if body.txn_ref is not None else doc.get('txn_ref')
    remarks = body.remarks if body.remarks is not None else doc.get('remarks')

    pc = PaymentCreate(
        school_id=doc['school_id'],
        student_id=doc['student_id'],
        items=[PaymentLineItem(**i) for i in doc.get('items', [])],
        discount=approved_discount,
        late_fee=doc.get('late_fee', 0.0),
        payment_mode=payment_mode,
        txn_ref=txn_ref,
        remarks=remarks,
    )

    # Preserve the ORIGINAL requester as the "collected_by" on the Payment doc
    # (so audit shows who initiated the collection flow), but stamp the
    # actual collector on the approval so both are visible.
    requester = await users_col.find_one({'id': doc.get('requested_by_id')}, {'_id': 0})
    acting_user = requester or current
    payment = await _finalize_payment(doc['school_id'], pc, acting_user)

    # Stamp the approval trail on the payment doc so audit/receipt shows it.
    await payments_col.update_one(
        {'id': payment.id},
        {'$set': {
            'discount_approval_id': approval_id,
            'discount_reason': doc.get('discount_reason'),
            'discount_approved_by_id': doc.get('reviewed_by_id'),
            'discount_approved_by_name': doc.get('reviewed_by_name'),
            'discount_approved_at': doc.get('reviewed_at'),
        }}
    )
    # Update the approval doc → collected.
    await discount_approvals_col.update_one(
        {'id': approval_id},
        {'$set': {
            'status': 'collected',
            'collected_by_id': current.get('id'),
            'collected_by_name': current.get('full_name'),
            'collected_at': now_iso(),
            'payment_id': payment.id,
            'receipt_number': payment.receipt_number,
            'payment_mode': payment_mode,
            'txn_ref': txn_ref,
            'remarks': remarks,
        }}
    )
    await log_audit(action='discount_approval.collect', current_user=current,
                    school_id=doc['school_id'], entity_type='discount_approval',
                    entity_id=approval_id,
                    details={'student': doc.get('student_name'),
                             'approved_discount': approved_discount,
                             'payment_id': payment.id,
                             'receipt': payment.receipt_number})
    return {
        'ok': True,
        'status': 'collected',
        'payment_id': payment.id,
        'receipt_number': payment.receipt_number,
        'payment': payment.model_dump(),
    }


@api.post('/discount-approvals/{approval_id}/reject',
          dependencies=[Depends(require_roles('owner'))])
async def reject_discount_request(approval_id: str, body: DiscountApprovalReview,
                                  current=Depends(get_current_user)):
    doc = await discount_approvals_col.find_one({'id': approval_id}, {'_id': 0})
    if not doc:
        raise HTTPException(404, 'Approval not found')
    if doc.get('status') != 'pending':
        raise HTTPException(400, f"Approval already {doc.get('status')}")
    await discount_approvals_col.update_one(
        {'id': approval_id},
        {'$set': {
            'status': 'rejected',
            'reviewed_by_id': current.get('id'),
            'reviewed_by_name': current.get('full_name'),
            'reviewed_at': now_iso(),
            'review_remark': (body.remark or '').strip() or None,
        }}
    )
    await log_audit(action='discount_approval.reject', current_user=current,
                    school_id=doc['school_id'], entity_type='discount_approval',
                    entity_id=approval_id,
                    details={'student': doc.get('student_name'),
                             'discount': doc.get('discount'),
                             'reason': body.remark})
    return {'ok': True, 'status': 'rejected'}


@api.post('/payments/razorpay/order')
async def razorpay_create_order(body: RazorpayOrderRequest, request: Request,
                                current=Depends(get_current_user)):
    student = await students_col.find_one({'id': body.student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    if current['role'] == 'parent' and not parent_can_access_student(current, body.student_id):
        raise HTTPException(403, 'Forbidden')
    # Discounts on online (razorpay) payments must also go through owner
    # approval — but the online rails complicate the money flow if we take the
    # payment first and reject later. So for online payments we simply reject
    # discount>0 and instruct the collector to use the offline flow.
    if float(body.discount or 0) > 0 and current['role'] in ('school_admin', 'accountant', 'super_admin'):
        raise HTTPException(400, 'Discounts on online (Razorpay) payments are not allowed. '
                                  'Please use the offline (cash/UPI/cheque/bank_transfer) mode so '
                                  'the discount can be sent to the school owner for approval before '
                                  'the receipt is generated.')
    subtotal = sum(i.amount for i in body.items)
    total = subtotal + body.late_fee - body.discount
    amount_paise = int(round(total * 100))
    if amount_paise <= 0:
        raise HTTPException(400, 'Amount must be > 0')
    receipt_ref = f'stv_{new_id()[:8]}'
    if RAZORPAY_MOCK:
        order = {'id': f'order_mock_{new_id()[:12]}'}
    else:
        order = rzp_client.order.create(data={
            'amount': amount_paise, 'currency': 'INR',
            'receipt': receipt_ref, 'payment_capture': 1,
            'notes': {
                'school_id': student['school_id'],
                'student_id': student['id'],
                'student_name': student['full_name'],
            }
        })
    await razorpay_orders_col.insert_one({
        'id': order['id'],
        'school_id': student['school_id'],
        'student_id': student['id'],
        'amount': total,
        'items': [i.model_dump() for i in body.items],
        'discount': body.discount,
        'late_fee': body.late_fee,
        'remarks': body.remarks,
        'status': 'created',
        'created_at': now_iso(),
    })
    return {
        'order_id': order['id'],
        'amount': amount_paise,
        'currency': 'INR',
        'key_id': RAZORPAY_KEY_ID or MOCK_KEY_ID,
        'student_name': student['full_name'],
        'mock': RAZORPAY_MOCK,
    }


@api.post('/payments/razorpay/verify')
async def razorpay_verify(body: RazorpayVerifyRequest, request: Request,
                          current=Depends(get_current_user)):
    # Verify signature (in mock mode we validate against MOCK_KEY_SECRET)
    secret = MOCK_KEY_SECRET if RAZORPAY_MOCK else RAZORPAY_KEY_SECRET
    payload = f"{body.razorpay_order_id}|{body.razorpay_payment_id}".encode()
    expected = hmac.new(secret.encode(), payload, hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, body.razorpay_signature):
        raise HTTPException(400, 'Signature verification failed')
    order = await razorpay_orders_col.find_one({'id': body.razorpay_order_id}, {'_id': 0})
    if not order:
        raise HTTPException(404, 'Order not found')
    if order['status'] == 'paid':
        # idempotent
        pay = await payments_col.find_one({'razorpay_order_id': body.razorpay_order_id}, {'_id': 0})
        return pay
    pc = PaymentCreate(
        school_id=order['school_id'],
        student_id=order['student_id'],
        items=[PaymentLineItem(**i) for i in order['items']],
        discount=order.get('discount', 0),
        late_fee=order.get('late_fee', 0),
        payment_mode='razorpay',
        txn_ref=body.razorpay_payment_id,
        remarks=order.get('remarks'),
    )
    payment = await _finalize_payment(order['school_id'], pc, current,
                                      razorpay_order_id=body.razorpay_order_id,
                                      razorpay_payment_id=body.razorpay_payment_id,
                                      razorpay_signature=body.razorpay_signature)
    await razorpay_orders_col.update_one({'id': body.razorpay_order_id},
                                         {'$set': {'status': 'paid', 'payment_id': payment.id}})
    return payment.model_dump()


@api.post('/payments/razorpay/webhook')
async def razorpay_webhook(request: Request):
    body = await request.body()
    sig = request.headers.get('X-Razorpay-Signature', '')
    secret = RAZORPAY_WEBHOOK_SECRET or RAZORPAY_KEY_SECRET
    expected = hmac.new(secret.encode(), body, hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, sig):
        raise HTTPException(400, 'Bad webhook signature')
    # Optionally handle events here
    return {'ok': True}


@api.get('/payments')
async def list_payments(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        student_id: Optional[str] = None,
                        start_date: Optional[str] = None,
                        end_date: Optional[str] = None,
                        mode: Optional[str] = None,
                        limit: int = 500):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        if not child_ids:
            return []
        q['student_id'] = {'$in': child_ids}
    if student_id:
        q['student_id'] = student_id
    if mode:
        q['payment_mode'] = mode
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date + 'T23:59:59'
        q['paid_at'] = rq
    rows = await payments_col.find(q, {'_id': 0}).sort('paid_at', -1).limit(limit).to_list(limit)
    return rows


@api.get('/payments/{payment_id}')
async def get_payment(payment_id: str, current=Depends(get_current_user)):
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if current['role'] != 'super_admin' and p['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    if current['role'] == 'parent' and not parent_can_access_student(current, p.get('student_id')):
        raise HTTPException(403, 'Forbidden')
    return p


@api.get('/payments/{payment_id}/receipt.pdf')
async def download_receipt(payment_id: str, current=Depends(get_current_user)):
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if current['role'] != 'super_admin' and p['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    if current['role'] == 'parent' and not parent_can_access_student(current, p.get('student_id')):
        raise HTTPException(403, 'Forbidden')
    school = await schools_col.find_one({'id': p['school_id']}, {'_id': 0}) or {}
    student = await students_col.find_one({'id': p['student_id']}, {'_id': 0}) or {}
    # enrich with class name
    if student.get('class_id'):
        cls = await classes_col.find_one({'id': student['class_id']}, {'_id': 0})
        if cls:
            student['class_name'] = cls['name']
    pdf_bytes = generate_receipt_pdf(p, school, student)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type='application/pdf',
        headers={'Content-Disposition': f'inline; filename="receipt_{p["receipt_number"]}.pdf"'}
    )


# ---------- SUPER ADMIN: EDIT / VOID / RESTORE RECEIPTS ----------
def _payment_snapshot(p: dict) -> dict:
    """Return an immutable snapshot of the mutable financial fields of a payment.
    Stored inside `edit_history` for a full audit trail of every super-admin edit."""
    return {
        'items': p.get('items') or [],
        'subtotal': p.get('subtotal'),
        'discount': p.get('discount'),
        'late_fee': p.get('late_fee'),
        'total_paid': p.get('total_paid'),
        'payment_mode': p.get('payment_mode'),
        'txn_ref': p.get('txn_ref'),
        'remarks': p.get('remarks'),
        'paid_at': p.get('paid_at'),
        'status': p.get('status'),
    }


@api.patch('/payments/{payment_id}',
           dependencies=[Depends(require_roles('super_admin'))])
async def edit_payment(payment_id: str, body: PaymentEdit,
                       current=Depends(get_current_user)):
    """Super-admin only. Edit a previously collected payment (fix wrong items,
    amounts, mode, etc.). The `receipt_number` is preserved for financial
    continuity; every change is recorded in `edit_history` for audit."""
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if not (body.reason and body.reason.strip()):
        raise HTTPException(400, 'Reason is required for editing a receipt')
    if p.get('status') == 'voided':
        raise HTTPException(400, 'Cannot edit a voided receipt. Restore it first.')

    # Snapshot BEFORE mutation.
    prev = _payment_snapshot(p)
    prev['edited_at'] = now_iso()
    prev['edited_by_id'] = current.get('id')
    prev['edited_by_name'] = current.get('full_name')
    prev['reason'] = body.reason.strip()

    update: Dict[str, Any] = {}
    if body.items is not None:
        items_dump = [i.model_dump() for i in body.items]
        subtotal = sum(float(i.get('amount') or 0) for i in items_dump)
        discount = float(body.discount if body.discount is not None else (p.get('discount') or 0))
        late_fee = float(body.late_fee if body.late_fee is not None else (p.get('late_fee') or 0))
        update['items'] = items_dump
        update['subtotal'] = subtotal
        update['discount'] = discount
        update['late_fee'] = late_fee
        update['total_paid'] = round(subtotal + late_fee - discount, 2)
    else:
        # items unchanged, but discount / late_fee may still change → recompute total
        if body.discount is not None or body.late_fee is not None:
            subtotal = float(p.get('subtotal') or 0)
            discount = float(body.discount if body.discount is not None else (p.get('discount') or 0))
            late_fee = float(body.late_fee if body.late_fee is not None else (p.get('late_fee') or 0))
            update['discount'] = discount
            update['late_fee'] = late_fee
            update['total_paid'] = round(subtotal + late_fee - discount, 2)

    if body.payment_mode is not None:
        update['payment_mode'] = body.payment_mode
    if body.txn_ref is not None:
        update['txn_ref'] = body.txn_ref
    if body.remarks is not None:
        update['remarks'] = body.remarks
    if body.paid_at is not None and body.paid_at.strip():
        update['paid_at'] = body.paid_at

    update['edited_at'] = prev['edited_at']
    update['edited_by_id'] = current.get('id')
    update['edited_by_name'] = current.get('full_name')
    update['edited_reason'] = body.reason.strip()

    await payments_col.update_one(
        {'id': payment_id},
        {'$set': update, '$push': {'edit_history': prev}},
    )

    await log_audit(action='payment.edit', current_user=current,
                    school_id=p.get('school_id'),
                    entity_type='payment', entity_id=payment_id,
                    details={
                        'receipt_number': p.get('receipt_number'),
                        'reason': body.reason.strip(),
                        'previous_total': prev.get('total_paid'),
                        'new_total': update.get('total_paid', prev.get('total_paid')),
                    })
    return await payments_col.find_one({'id': payment_id}, {'_id': 0})


@api.post('/payments/{payment_id}/void',
          dependencies=[Depends(require_roles('super_admin'))])
async def void_payment(payment_id: str, body: PaymentVoid,
                       current=Depends(get_current_user)):
    """Super-admin only. Void / cancel a receipt. The payment is retained for
    audit but its `status` is set to `voided`, which automatically excludes it
    from all fee-schedule / dues / collection-report aggregations."""
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if p.get('status') == 'voided':
        raise HTTPException(400, 'Receipt is already voided')
    if not (body.reason and body.reason.strip()):
        raise HTTPException(400, 'Reason is required to void a receipt')

    ts = now_iso()
    await payments_col.update_one(
        {'id': payment_id},
        {'$set': {
            'status': 'voided',
            'voided_at': ts,
            'voided_by_id': current.get('id'),
            'voided_by_name': current.get('full_name'),
            'void_reason': body.reason.strip(),
        }},
    )
    await log_audit(action='payment.void', current_user=current,
                    school_id=p.get('school_id'),
                    entity_type='payment', entity_id=payment_id,
                    details={
                        'receipt_number': p.get('receipt_number'),
                        'reason': body.reason.strip(),
                        'amount_reversed': p.get('total_paid'),
                        'student_id': p.get('student_id'),
                    })
    return await payments_col.find_one({'id': payment_id}, {'_id': 0})


@api.post('/payments/{payment_id}/restore',
          dependencies=[Depends(require_roles('super_admin'))])
async def restore_payment(payment_id: str, current=Depends(get_current_user)):
    """Super-admin only. Un-void a previously voided receipt (mistake reversal)."""
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if p.get('status') != 'voided':
        raise HTTPException(400, 'Receipt is not voided')

    await payments_col.update_one(
        {'id': payment_id},
        {'$set': {'status': 'success'},
         '$unset': {'voided_at': '', 'voided_by_id': '',
                    'voided_by_name': '', 'void_reason': ''}},
    )
    await log_audit(action='payment.restore', current_user=current,
                    school_id=p.get('school_id'),
                    entity_type='payment', entity_id=payment_id,
                    details={'receipt_number': p.get('receipt_number')})
    return await payments_col.find_one({'id': payment_id}, {'_id': 0})


# =====================================================
# ATTENDANCE
# =====================================================
@api.post('/attendance/mark', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'teacher'))])
async def mark_attendance(body: AttendanceBulkMark, request: Request,
                          current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    # Delete existing for that date/class first (upsert semantics)
    await attendance_col.delete_many({
        'school_id': sid,
        'date': body.date,
        'class_id': body.class_id,
        'section': body.section,
    })
    docs = []
    for e in body.entries:
        rec = AttendanceRecord(
            school_id=sid, date=body.date, class_id=body.class_id,
            section=body.section, student_id=e.get('student_id'),
            status=e.get('status', 'present'),
            remarks=e.get('remarks'),
            marked_by_id=current.get('id'),
        )
        docs.append(rec.model_dump())
    if docs:
        await attendance_col.insert_many(docs)
    await log_audit(action='attendance.mark', current_user=current, school_id=sid,
                    entity_type='attendance', entity_id=f"{body.date}-{body.class_id}",
                    details={'count': len(docs), 'date': body.date})
    return {'saved': len(docs)}


@api.get('/attendance')
async def list_attendance(request: Request, current=Depends(get_current_user),
                          school_id: Optional[str] = None,
                          date: Optional[str] = None,
                          class_id: Optional[str] = None,
                          section: Optional[str] = None,
                          student_id: Optional[str] = None,
                          start_date: Optional[str] = None,
                          end_date: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if date:
        q['date'] = date
    if class_id:
        q['class_id'] = class_id
    if section:
        q['section'] = section
    if student_id:
        q['student_id'] = student_id
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        if not child_ids:
            return []
        q['student_id'] = {'$in': child_ids}
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date
        q['date'] = rq if rq else q.get('date')
    rows = await attendance_col.find(q, {'_id': 0}).limit(5000).to_list(5000)
    return rows


# =====================================================
# HOMEWORK
# =====================================================
@api.get('/homework')
async def list_homework(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        class_id: Optional[str] = None,
                        section: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if class_id:
        q['class_id'] = class_id
    if section:
        q['section'] = section
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        if not child_ids:
            return []
        # Homework is class-scoped; pull the child's classes and match
        children = await students_col.find(
            {'id': {'$in': child_ids}}, {'_id': 0, 'class_id': 1, 'section': 1}
        ).to_list(100)
        class_ids = list({c.get('class_id') for c in children if c.get('class_id')})
        if not class_ids:
            return []
        q['class_id'] = {'$in': class_ids}
    rows = await homework_col.find(q, {'_id': 0}).sort('created_at', -1).limit(500).to_list(500)
    return rows


@api.post('/homework', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'teacher'))])
async def create_homework(body: HomeworkCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    hw = Homework(school_id=sid, class_id=body.class_id, section=body.section,
                  subject=body.subject, title=body.title, description=body.description,
                  due_date=body.due_date, attachment_url=body.attachment_url,
                  created_by_id=current.get('id'),
                  created_by_name=current.get('full_name'))
    await homework_col.insert_one(hw.model_dump())
    await log_audit(action='homework.create', current_user=current, school_id=sid,
                    entity_type='homework', entity_id=hw.id, details={'title': hw.title})
    return hw.model_dump()


# =====================================================
# TIMETABLE
# =====================================================
@api.get('/timetable')
async def get_timetable(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        class_id: Optional[str] = None,
                        section: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if class_id:
        q['class_id'] = class_id
    if section:
        q['section'] = section
    return await timetable_col.find(q, {'_id': 0}).to_list(200)


@api.post('/timetable', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def upsert_timetable(body: TimetableCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    exists = await timetable_col.find_one({'school_id': sid, 'class_id': body.class_id, 'section': body.section}, {'_id': 0})
    slots = [s.model_dump() for s in body.slots]
    if exists:
        await timetable_col.update_one({'id': exists['id']}, {'$set': {'slots': slots, 'updated_at': now_iso()}})
        return await timetable_col.find_one({'id': exists['id']}, {'_id': 0})
    t = Timetable(school_id=sid, class_id=body.class_id, section=body.section, slots=body.slots)
    await timetable_col.insert_one(t.model_dump())
    return t.model_dump()


# =====================================================
# EVENTS, CIRCULARS, GALLERY, STAFF, NOTIFICATIONS
# =====================================================
@api.get('/events')
async def list_events(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await events_col.find({'school_id': sid}, {'_id': 0}).sort('event_date', -1).limit(200).to_list(200)


@api.post('/events', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_event(body: EventCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    e = Event(school_id=sid, title=body.title, description=body.description,
              event_date=body.event_date, location=body.location, image_url=body.image_url)
    await events_col.insert_one(e.model_dump())
    await log_audit(action='event.create', current_user=current, school_id=sid,
                    entity_type='event', entity_id=e.id, details={'title': e.title})
    return e.model_dump()


@api.get('/circulars')
async def list_circulars(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await circulars_col.find({'school_id': sid, 'status': 'published'}, {'_id': 0}).sort('created_at', -1).limit(200).to_list(200)


@api.post('/circulars', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_circular(body: CircularCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    c = Circular(school_id=sid, title=body.title, body=body.body, priority=body.priority,
                 status=body.status, publish_at=body.publish_at,
                 attachment_url=body.attachment_url, audience=body.audience,
                 class_id=body.class_id, created_by_name=current.get('full_name'))
    await circulars_col.insert_one(c.model_dump())
    await log_audit(action='circular.create', current_user=current, school_id=sid,
                    entity_type='circular', entity_id=c.id, details={'title': c.title})
    return c.model_dump()


@api.get('/gallery')
async def list_gallery(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await gallery_col.find({'school_id': sid}, {'_id': 0}).sort('created_at', -1).to_list(200)


@api.post('/gallery', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_album(body: GalleryAlbumCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    a = GalleryAlbum(school_id=sid, title=body.title, description=body.description,
                     cover_url=body.cover_url, photos=body.photos)
    await gallery_col.insert_one(a.model_dump())
    await log_audit(action='gallery.create', current_user=current, school_id=sid,
                    entity_type='gallery', entity_id=a.id, details={'title': a.title})
    return a.model_dump()


@api.get('/staff')
async def list_staff(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await staff_col.find({'school_id': sid}, {'_id': 0}).to_list(500)


@api.post('/staff', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_staff(body: StaffCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    s = Staff(school_id=sid, full_name=body.full_name, email=body.email, phone=body.phone,
              designation=body.designation, department=body.department,
              subjects=body.subjects, joining_date=body.joining_date, photo_url=body.photo_url)
    await staff_col.insert_one(s.model_dump())
    await log_audit(action='staff.create', current_user=current, school_id=sid,
                    entity_type='staff', entity_id=s.id, details={'name': s.full_name})
    return s.model_dump()


@api.get('/notifications')
async def list_notifications(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    # audience filter for role-scoped users
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        q['$or'] = [{'audience': 'all'}, {'audience': 'parents'},
                    {'student_ids': {'$in': child_ids}} if child_ids else {'student_ids': None}]
    elif current['role'] == 'teacher':
        q['$or'] = [{'audience': 'all'}, {'audience': 'teachers'}]
    return await notifications_col.find(q, {'_id': 0}).sort('created_at', -1).limit(200).to_list(200)


@api.post('/notifications', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'teacher'))])
async def create_notification(body: NotificationCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    n = Notification(school_id=sid, title=body.title, body=body.body,
                     audience=body.audience, class_id=body.class_id,
                     student_ids=body.student_ids, kind=body.kind)
    await notifications_col.insert_one(n.model_dump())
    await log_audit(action='notification.create', current_user=current, school_id=sid,
                    entity_type='notification', entity_id=n.id,
                    details={'title': n.title, 'audience': n.audience})
    return n.model_dump()


# =====================================================
# REPORTS
# =====================================================
@api.get('/reports/collection')
async def report_collection(request: Request, current=Depends(get_current_user),
                            school_id: Optional[str] = None,
                            start_date: Optional[str] = None,
                            end_date: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid, 'status': 'success'}
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date + 'T23:59:59'
        q['paid_at'] = rq
    payments = await payments_col.find(q, {'_id': 0}).sort('paid_at', -1).to_list(5000)
    total = sum(p.get('total_paid', 0) for p in payments)
    by_mode: Dict[str, float] = {}
    for p in payments:
        m = p.get('payment_mode', 'cash')
        by_mode[m] = by_mode.get(m, 0) + p.get('total_paid', 0)
    return {'payments': payments, 'total': total, 'by_mode': by_mode, 'count': len(payments)}


@api.get('/reports/collection.pdf')
async def report_collection_pdf(request: Request, current=Depends(get_current_user),
                                school_id: Optional[str] = None,
                                start_date: Optional[str] = None,
                                end_date: Optional[str] = None):
    data = await report_collection(request, current, school_id, start_date, end_date)
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    school = await schools_col.find_one({'id': sid}, {'_id': 0}) or {}
    cols = ['Receipt No', 'Date', 'Student', 'Mode', 'Amount (Rs.)']
    rows = []
    for p in data['payments']:
        rows.append([p.get('receipt_number', ''),
                     (p.get('paid_at') or '')[:10],
                     p.get('student_name', ''),
                     str(p.get('payment_mode', '')).replace('_', ' ').title(),
                     f"{p.get('total_paid', 0):,.2f}"])
    subtitle = f"Period: {start_date or 'All'}  to  {end_date or 'Today'}"
    summary = {'Total Collection': f"Rs. {data['total']:,.2f}", 'Transactions': data['count']}
    pdf = generate_report_pdf('Fee Collection Report', subtitle, cols, rows,
                              school.get('name', 'Stanvard School'), summary)
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf',
                             headers={'Content-Disposition': 'inline; filename="collection_report.pdf"'})


@api.get('/reports/collection.csv')
async def report_collection_csv(request: Request, current=Depends(get_current_user),
                                school_id: Optional[str] = None,
                                start_date: Optional[str] = None,
                                end_date: Optional[str] = None):
    data = await report_collection(request, current, school_id, start_date, end_date)
    import csv
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Receipt No', 'Date', 'Student', 'Mode', 'Amount'])
    for p in data['payments']:
        writer.writerow([p.get('receipt_number', ''),
                        (p.get('paid_at') or '')[:10],
                        p.get('student_name', ''),
                        p.get('payment_mode', ''),
                        p.get('total_paid', 0)])
    return StreamingResponse(io.BytesIO(buffer.getvalue().encode()),
                             media_type='text/csv',
                             headers={'Content-Disposition': 'attachment; filename="collection.csv"'})


@api.get('/reports/collection.xlsx')
async def report_collection_xlsx(request: Request, current=Depends(get_current_user),
                                 school_id: Optional[str] = None,
                                 start_date: Optional[str] = None,
                                 end_date: Optional[str] = None):
    from openpyxl import Workbook
    data = await report_collection(request, current, school_id, start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Collection'
    ws.append(['Receipt No', 'Date', 'Student', 'Mode', 'Amount'])
    for p in data['payments']:
        ws.append([p.get('receipt_number', ''),
                   (p.get('paid_at') or '')[:10],
                   p.get('student_name', ''),
                   p.get('payment_mode', ''),
                   p.get('total_paid', 0)])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': 'attachment; filename="collection.xlsx"'})


@api.get('/reports/attendance')
async def report_attendance(request: Request, current=Depends(get_current_user),
                            school_id: Optional[str] = None,
                            start_date: Optional[str] = None,
                            end_date: Optional[str] = None,
                            class_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date
        q['date'] = rq
    if class_id:
        q['class_id'] = class_id
    rows = await attendance_col.find(q, {'_id': 0}).to_list(20000)
    total = len(rows)
    present = sum(1 for r in rows if r.get('status') == 'present')
    absent = sum(1 for r in rows if r.get('status') == 'absent')
    leave = sum(1 for r in rows if r.get('status') == 'leave')
    return {'total': total, 'present': present, 'absent': absent, 'leave': leave,
            'attendance_rate': round(present / total * 100, 2) if total else 0.0}


# =====================================================
# DASHBOARD
# =====================================================
@api.get('/dashboard/summary')
async def dashboard_summary(request: Request, current=Depends(get_current_user),
                            school_id: Optional[str] = None):
    sid = await resolve_school_id_safe(current, school_id, request.headers.get('X-School-Id'))
    today = datetime.now().strftime('%Y-%m-%d')
    month_start = datetime.now().strftime('%Y-%m-01')

    total_students = await students_col.count_documents({'school_id': sid, 'status': 'active'})
    total_staff = await staff_col.count_documents({'school_id': sid, 'status': 'active'})

    # Today's collection (exclude voided receipts)
    today_payments = await payments_col.find({'school_id': sid, 'status': 'success', 'paid_at': {'$gte': today, '$lt': today + 'T23:59:59'}}, {'_id': 0}).to_list(2000)
    today_collection = sum(p.get('total_paid', 0) for p in today_payments)
    # Monthly collection (exclude voided receipts)
    month_payments = await payments_col.find({'school_id': sid, 'status': 'success', 'paid_at': {'$gte': month_start}}, {'_id': 0}).to_list(5000)
    monthly_collection = sum(p.get('total_paid', 0) for p in month_payments)

    # Attendance today
    today_att = await attendance_col.find({'school_id': sid, 'date': today}, {'_id': 0}).to_list(5000)
    present_today = sum(1 for a in today_att if a.get('status') == 'present')
    absent_today = sum(1 for a in today_att if a.get('status') == 'absent')

    # New admissions (this month)
    new_admissions = await students_col.count_documents({
        'school_id': sid,
        'created_at': {'$gte': month_start}
    })

    # Recent activity
    recent_payments = await payments_col.find({'school_id': sid}, {'_id': 0}).sort('paid_at', -1).limit(5).to_list(5)
    upcoming_events = await events_col.find({
        'school_id': sid, 'event_date': {'$gte': today}
    }, {'_id': 0}).sort('event_date', 1).limit(5).to_list(5)
    recent_circulars = await circulars_col.find({
        'school_id': sid, 'status': 'published'
    }, {'_id': 0}).sort('created_at', -1).limit(5).to_list(5)

    # Pending fees (naive: students with assignments who haven't paid this month) — exclude voided
    with_assignments = await fee_assignments_col.distinct('student_id', {'school_id': sid})
    paid_this_month = await payments_col.distinct('student_id', {'school_id': sid, 'status': 'success', 'paid_at': {'$gte': month_start}})
    pending_students = max(len(with_assignments) - len(paid_this_month), 0)

    # Collection trend last 7 days (exclude voided)
    trend = []
    for i in range(6, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        payments = await payments_col.find({'school_id': sid, 'status': 'success', 'paid_at': {'$gte': d, '$lt': d + 'T23:59:59'}}, {'_id': 0}).to_list(1000)
        trend.append({'date': d, 'amount': sum(p.get('total_paid', 0) for p in payments)})

    return {
        'total_students': total_students,
        'total_staff': total_staff,
        'today_collection': today_collection,
        'monthly_collection': monthly_collection,
        'present_today': present_today,
        'absent_today': absent_today,
        'new_admissions': new_admissions,
        'pending_students': pending_students,
        'recent_payments': recent_payments,
        'upcoming_events': upcoming_events,
        'recent_circulars': recent_circulars,
        'collection_trend': trend,
    }


# =====================================================
# AUDIT LOGS
# =====================================================
@api.get('/audit-logs', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def list_audit(request: Request, current=Depends(get_current_user),
                    school_id: Optional[str] = None, limit: int = 200):
    q: Dict[str, Any] = {}
    if current['role'] == 'super_admin':
        if school_id:
            q['school_id'] = school_id
    else:
        q['school_id'] = current['school_id']
    rows = await audit_col.find(q, {'_id': 0}).sort('created_at', -1).limit(limit).to_list(limit)
    return rows


# =====================================================
# SETTINGS
# =====================================================
@api.get('/settings')
async def get_settings(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    s = await settings_col.find_one({'school_id': sid}, {'_id': 0})
    if not s:
        s = SchoolSettings(school_id=sid).model_dump()
        await settings_col.insert_one(dict(s))
        s = await settings_col.find_one({'school_id': sid}, {'_id': 0})
    return s


@api.patch('/settings', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_settings(payload: Dict[str, Any], request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, None, request.headers.get('X-School-Id'))
    payload['updated_at'] = now_iso()
    await settings_col.update_one({'school_id': sid}, {'$set': payload}, upsert=True)
    await log_audit(action='settings.update', current_user=current, school_id=sid,
                    entity_type='settings', details=payload)
    return await settings_col.find_one({'school_id': sid}, {'_id': 0})


# =====================================================
# USER DELETE + PASSWORD RESET
# =====================================================
@api.delete('/users/{user_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_user(user_id: str, current=Depends(get_current_user)):
    target = await users_col.find_one({'id': user_id}, {'_id': 0})
    if not target:
        raise HTTPException(404, 'User not found')
    if current['role'] == 'school_admin':
        if target.get('school_id') != current.get('school_id') or target.get('role') == 'super_admin':
            raise HTTPException(403, 'Cannot delete this user')
    if target['id'] == current['id']:
        raise HTTPException(400, 'Cannot delete yourself')
    await users_col.update_one({'id': user_id}, {'$set': {'status': 'inactive', 'updated_at': now_iso()}})
    await log_audit(action='user.delete', current_user=current,
                    entity_type='user', entity_id=user_id, details={'email': target.get('email')})
    return {'ok': True}


@api.post('/users/{user_id}/reset-password', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def reset_password(user_id: str, payload: Dict[str, str], current=Depends(get_current_user)):
    new_pw = payload.get('password')
    if not new_pw or len(new_pw) < 6:
        raise HTTPException(400, 'Password must be at least 6 characters')
    target = await users_col.find_one({'id': user_id}, {'_id': 0})
    if not target:
        raise HTTPException(404, 'User not found')
    if current['role'] == 'school_admin' and target.get('school_id') != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    await users_col.update_one({'id': user_id}, {'$set': {'password_hash': hash_password(new_pw), 'updated_at': now_iso()}})
    await log_audit(action='user.reset_password', current_user=current,
                    entity_type='user', entity_id=user_id, details={'email': target.get('email')})
    return {'ok': True}


# =====================================================
# ANALYTICS
# =====================================================
@api.get('/analytics')
async def analytics(request: Request, current=Depends(get_current_user),
                   school_id: Optional[str] = None, year: Optional[int] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    year = year or datetime.now().year
    year_start = f'{year}-01-01'
    year_end = f'{year}-12-31T23:59:59'

    # All payments for the year
    payments = await payments_col.find({
        'school_id': sid, 'status': 'success',
        'paid_at': {'$gte': year_start, '$lte': year_end}
    }, {'_id': 0}).to_list(20000)

    # Monthly breakdown (Jan..Dec)
    months = [{'month': datetime(year, m, 1).strftime('%b'),
               'received': 0.0, 'transactions': 0, 'discount': 0.0, 'late_fee': 0.0}
              for m in range(1, 13)]
    by_mode = {}
    by_head = {}
    for p in payments:
        try:
            paid_month = int((p.get('paid_at') or '')[5:7])
            months[paid_month - 1]['received'] += p.get('total_paid', 0)
            months[paid_month - 1]['transactions'] += 1
            months[paid_month - 1]['discount'] += p.get('discount', 0)
            months[paid_month - 1]['late_fee'] += p.get('late_fee', 0)
        except Exception:
            pass
        mode = p.get('payment_mode', 'cash')
        by_mode[mode] = by_mode.get(mode, 0) + p.get('total_paid', 0)
        for item in p.get('items', []):
            hd = item.get('fee_head_name', 'Other')
            by_head[hd] = by_head.get(hd, 0) + item.get('amount', 0)

    total_received = sum(p.get('total_paid', 0) for p in payments)
    total_transactions = len(payments)
    total_discount = sum(p.get('discount', 0) for p in payments)
    total_late_fee = sum(p.get('late_fee', 0) for p in payments)

    # Expected: sum of all assignment amounts (custom_items or plan)
    assignments = await fee_assignments_col.find({'school_id': sid}, {'_id': 0}).to_list(10000)
    total_expected = 0.0
    plans_cache: Dict[str, Any] = {}
    for a in assignments:
        items = a.get('custom_items') or []
        if not items and a.get('fee_plan_id'):
            pid = a['fee_plan_id']
            if pid not in plans_cache:
                plans_cache[pid] = await fee_plans_col.find_one({'id': pid}, {'_id': 0}) or {}
            items = plans_cache[pid].get('items', [])
        for it in items:
            total_expected += it.get('amount', 0)
        total_expected -= a.get('discount_amount', 0)
    total_due = max(total_expected - total_received, 0)

    # Attendance summary for the year
    att = await attendance_col.find({'school_id': sid, 'date': {'$gte': year_start[:10], '$lte': year_end[:10]}}, {'_id': 0}).to_list(50000)
    att_total = len(att)
    att_present = sum(1 for a in att if a.get('status') == 'present')
    att_absent = sum(1 for a in att if a.get('status') == 'absent')
    att_leave = sum(1 for a in att if a.get('status') == 'leave')

    # New admissions per month
    students = await students_col.find({'school_id': sid}, {'_id': 0}).to_list(5000)
    adm_months = [0] * 12
    for s in students:
        try:
            ca = s.get('created_at') or s.get('admission_date') or ''
            if ca and int(ca[:4]) == year:
                adm_months[int(ca[5:7]) - 1] += 1
        except Exception:
            pass
    for i, m in enumerate(months):
        m['admissions'] = adm_months[i]

    # Class-wise pending (approximation)
    classes = await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(200)
    by_class = []
    for c in classes:
        students_in_class = [s for s in students if s.get('class_id') == c['id']]
        payer_ids = {p['student_id'] for p in payments if p['student_id'] in {s['id'] for s in students_in_class}}
        by_class.append({
            'class_id': c['id'], 'class_name': c['name'],
            'total_students': len(students_in_class),
            'paying_students': len(payer_ids),
            'pending_students': max(len(students_in_class) - len(payer_ids), 0),
        })

    return {
        'year': year,
        'total_received': total_received,
        'total_expected': total_expected,
        'total_due': total_due,
        'total_discount': total_discount,
        'total_late_fee': total_late_fee,
        'total_transactions': total_transactions,
        'months': months,
        'by_mode': by_mode,
        'by_head': by_head,
        'by_class': by_class,
        'attendance': {
            'total': att_total,
            'present': att_present,
            'absent': att_absent,
            'leave': att_leave,
            'rate': round(att_present / att_total * 100, 2) if att_total else 0.0,
        },
        'total_students': len(students),
    }


# =====================================================
# ENHANCED REPORTS - Fee Status per student (paid/pending)
# =====================================================
@api.get('/reports/fee-status', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant', 'owner'))])
async def report_fee_status(request: Request, current=Depends(get_current_user),
                            school_id: Optional[str] = None,
                            class_id: Optional[str] = None,
                            section: Optional[str] = None,
                            class_sections: Optional[str] = None,
                            status_filter: Optional[str] = None,   # paid|partial|unpaid
                            due_min: Optional[float] = None,
                            due_max: Optional[float] = None,
                            payment_date_start: Optional[str] = None,
                            payment_date_end: Optional[str] = None,
                            quick_view: Optional[str] = None,      # defaulters|fully_paid|upcoming
                            behavior: Optional[str] = None,        # regular|late|defaulter
                            as_of_date: Optional[str] = None):     # YYYY-MM-DD — snapshot date
    """Student fee status report.

    Supports both:
      - Legacy single-filter: `class_id` + `section` (kept for backwards compatibility)
      - New multi-select: `class_sections` = comma-separated pairs like
        "<class_id>:<section>,<class_id>:,<class_id>:A"
        A blank section (e.g. "abc:") means "all sections of that class".

    Extended fields per row:
      - last_payment_date (ISO date str or None)
      - overdue_days (int, 0 if not overdue)
      - behavior_tag ('regular' | 'late' | 'defaulter' | 'na')
      - upcoming_due_date (nearest future due date, if any)
      - collection_percent
    Plus class/section rollups in `by_class`.
    """
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))

    cs_pairs: List[Dict[str, Optional[str]]] = []
    if class_sections:
        for token in class_sections.split(','):
            token = token.strip()
            if not token:
                continue
            if ':' in token:
                cid, sec = token.split(':', 1)
                cs_pairs.append({'class_id': cid.strip(), 'section': (sec.strip() or None)})
            else:
                cs_pairs.append({'class_id': token, 'section': None})

    student_q: Dict[str, Any] = {'school_id': sid, 'status': 'active'}
    if cs_pairs:
        or_clauses: List[Dict[str, Any]] = []
        for p in cs_pairs:
            clause: Dict[str, Any] = {'class_id': p['class_id']}
            if p['section']:
                clause['section'] = p['section']
            or_clauses.append(clause)
        student_q['$or'] = or_clauses
    else:
        if class_id:
            student_q['class_id'] = class_id
        if section:
            student_q['section'] = section
    students = await students_col.find(student_q, {'_id': 0}).to_list(5000)

    classes = await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(200)
    class_map = {c['id']: c['name'] for c in classes}

    student_ids = [s['id'] for s in students]
    assignments = await fee_assignments_col.find({'student_id': {'$in': student_ids}}, {'_id': 0}).to_list(20000)
    payments = await payments_col.find({'student_id': {'$in': student_ids}, 'status': 'success'}, {'_id': 0}).to_list(20000)

    plan_cache: Dict[str, Any] = {}
    async def _get_plan(pid):
        if pid not in plan_cache:
            plan_cache[pid] = await fee_plans_col.find_one({'id': pid}, {'_id': 0}) or {}
        return plan_cache[pid]

    expected_by_stu: Dict[str, float] = {}
    discount_by_stu: Dict[str, float] = {}
    due_date_by_stu: Dict[str, Optional[str]] = {}
    assignments_by_stu: Dict[str, list] = {}
    for a in assignments:
        sid_ = a['student_id']
        assignments_by_stu.setdefault(sid_, []).append(a)
        items = a.get('custom_items') or []
        if not items and a.get('fee_plan_id'):
            plan = await _get_plan(a['fee_plan_id'])
            items = plan.get('items', [])
        expected_by_stu[sid_] = expected_by_stu.get(sid_, 0.0) + sum(i.get('amount', 0) for i in items)
        discount_by_stu[sid_] = discount_by_stu.get(sid_, 0.0) + a.get('discount_amount', 0)
        if a.get('due_date') and not due_date_by_stu.get(sid_):
            due_date_by_stu[sid_] = a['due_date']

    # Payment aggregates: total paid, last date, late flag, per-mode counts
    paid_by_stu: Dict[str, float] = {}
    pay_disc_by_stu: Dict[str, float] = {}
    last_paid_by_stu: Dict[str, Optional[str]] = {}
    late_flag_by_stu: Dict[str, bool] = {}
    for p in payments:
        sid_ = p['student_id']
        paid_by_stu[sid_] = paid_by_stu.get(sid_, 0.0) + p.get('total_paid', 0)
        pay_disc_by_stu[sid_] = pay_disc_by_stu.get(sid_, 0.0) + float(p.get('discount') or 0)
        pd = (p.get('paid_at') or '')[:10]
        if pd and (not last_paid_by_stu.get(sid_) or pd > last_paid_by_stu[sid_]):
            last_paid_by_stu[sid_] = pd
        # A payment is late if made after the student's due_date.
        dd = due_date_by_stu.get(sid_)
        if dd and pd and pd > dd:
            late_flag_by_stu[sid_] = True

    # For monthly schedule computation we need per-student payments grouped
    payments_by_stu: Dict[str, list] = {}
    for p in payments:
        payments_by_stu.setdefault(p['student_id'], []).append(p)

    # Determine the current academic session — pick the most common one across
    # student assignments (fall back to 2026-27 if none).
    session_counts: Dict[str, int] = {}
    for a in assignments:
        s = a.get('academic_session')
        if s:
            session_counts[s] = session_counts.get(s, 0) + 1
    current_session = max(session_counts, key=session_counts.get) if session_counts else '2026-27'
    # If an as_of_date is supplied, all "today"-based computations (overdue,
    # upcoming, monthly status) use that reference date instead of the real
    # current date. Handy for retro/predictive Due List exports.
    try:
        today_dt = datetime.fromisoformat(as_of_date) if as_of_date else datetime.now()
    except Exception:
        today_dt = datetime.now()

    today = today_dt.date().isoformat()

    def _behavior_tag(paid: float, due: float, overdue_days: int, was_late: bool, expected: float) -> str:
        if expected <= 0:
            return 'na'
        if due <= 0 and paid > 0:
            return 'late' if was_late else 'regular'
        # Still owes something
        if overdue_days > 30:
            return 'defaulter'
        if was_late or overdue_days > 0:
            return 'late'
        return 'regular'

    rows = []
    for s in students:
        expected = expected_by_stu.get(s['id'], 0.0)
        disc = discount_by_stu.get(s['id'], 0.0)
        pay_disc = pay_disc_by_stu.get(s['id'], 0.0)
        # Approved per-payment discounts reduce what the student owes, same as
        # the assignment-level concession.
        expected_after_disc = max(expected - disc - pay_disc, 0.0)
        paid = paid_by_stu.get(s['id'], 0.0)
        due = max(expected_after_disc - paid, 0.0)
        row_status = 'unpaid' if paid == 0 else ('paid' if due <= 0 else 'partial')
        last_paid = last_paid_by_stu.get(s['id'])
        dd = due_date_by_stu.get(s['id'])
        overdue_days = 0
        upcoming_due = None
        if dd:
            try:
                if dd < today and due > 0:
                    overdue_days = (datetime.fromisoformat(today) - datetime.fromisoformat(dd)).days
                elif dd >= today:
                    upcoming_due = dd
            except Exception:
                overdue_days = 0
        was_late = late_flag_by_stu.get(s['id'], False)
        behavior_tag = _behavior_tag(paid, due, overdue_days, was_late, expected_after_disc)
        collection_pct = round((paid / expected_after_disc) * 100, 1) if expected_after_disc > 0 else 0.0

        # Filter application
        if status_filter and status_filter != 'all' and status_filter != row_status:
            continue
        if due_min is not None and due < due_min:
            continue
        if due_max is not None and due > due_max:
            continue
        if payment_date_start and (not last_paid or last_paid < payment_date_start):
            continue
        if payment_date_end and (not last_paid or last_paid > payment_date_end):
            continue
        if behavior and behavior != 'all' and behavior_tag != behavior:
            continue
        if quick_view == 'defaulters' and behavior_tag != 'defaulter':
            continue
        if quick_view == 'fully_paid' and row_status != 'paid':
            continue
        if quick_view == 'upcoming' and not upcoming_due:
            continue

        rows.append({
            'student_id': s['id'],
            'admission_number': s.get('admission_number'),
            'full_name': s.get('full_name'),
            'class_id': s.get('class_id'),
            'class_name': class_map.get(s.get('class_id'), '-'),
            'section': s.get('section'),
            'phone': s.get('phone'),
            'father_name': s.get('father_name'),
            'expected': round(expected_after_disc, 2),
            'gross_expected': round(expected, 2),
            'discount': round(disc + pay_disc, 2),
            'paid': round(paid, 2),
            'due': round(due, 2),
            'collection_percent': collection_pct,
            'due_date': dd,
            'upcoming_due_date': upcoming_due,
            'last_payment_date': last_paid,
            'overdue_days': overdue_days,
            'status': row_status,
            'behavior_tag': behavior_tag,
        })

    # ---- Compute monthly schedule for every student in the filtered result ----
    for r in rows:
        sid_ = r['student_id']
        sched, monthly_amt, fully_paid, oc, oa, _due = _build_month_schedule(
            r['expected'], payments_by_stu.get(sid_, []), current_session, today_dt,
            installments=_pick_assignment_installments(assignments_by_stu.get(sid_, [])),
            extra_paid_credit=pay_disc_by_stu.get(sid_, 0.0),
        )
        # Compact month tokens: only the fields the UI needs.
        r['monthly_amount'] = monthly_amt
        r['monthly_status'] = [
            {'i': e['index'], 'label': e['label'], 'status': e['status'],
             'paid': round(e['paid_amount'], 2), 'due': round(max(e['amount'] - e['paid_amount'], 0), 2)}
            for e in sched
        ]
        r['fully_paid'] = fully_paid
        r['overdue_months'] = oc
        r['overdue_amount'] = oa
        # "Due till date" — the amount that SHOULD have been paid by today for
        # a monthly-paying parent (past months + current month once past the
        # 15th grace day). Alias of overdue_amount for clearer UI labelling.
        r['due_till_date'] = oa

    total_expected = sum(r['expected'] for r in rows)
    total_paid = sum(r['paid'] for r in rows)
    total_due = sum(r['due'] for r in rows)
    total_discount = sum(r['discount'] for r in rows)
    total_gross = sum(r['gross_expected'] for r in rows)
    collection_pct = round((total_paid / total_expected) * 100, 1) if total_expected > 0 else 0.0

    # Class/Section rollups — monthly-status oriented
    by_class_map: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        key = f"{r['class_id']}::{r.get('section') or '-'}"
        b = by_class_map.setdefault(key, {
            'class_id': r['class_id'], 'class_name': r['class_name'],
            'section': r.get('section') or '-',
            'students': 0, 'expected': 0.0, 'paid': 0.0, 'due': 0.0,
            'paid_count': 0, 'partial_count': 0, 'unpaid_count': 0,
            'fully_paid_count': 0, 'students_with_dues': 0,
            'overdue_amount': 0.0,
        })
        b['students'] += 1
        b['expected'] += r['expected']
        b['paid'] += r['paid']
        b['due'] += r['due']
        b[f"{r['status']}_count"] += 1
        if r.get('fully_paid'):
            b['fully_paid_count'] += 1
        if r['due'] > 0:
            b['students_with_dues'] += 1
        b['overdue_amount'] += r.get('overdue_amount', 0.0)
    by_class = sorted(by_class_map.values(), key=lambda x: (x['class_name'], x['section']))
    for b in by_class:
        b['collection_percent'] = round((b['paid'] / b['expected']) * 100, 1) if b['expected'] > 0 else 0.0
        b['overdue_amount'] = round(b['overdue_amount'], 2)

    return {
        'rows': rows,
        'count': len(rows),
        'by_class': by_class,
        'academic_session': current_session,
        'summary': {
            'total_expected': total_expected,
            'total_gross_expected': total_gross,
            'total_paid': total_paid,
            'total_due': total_due,
            'total_discount': total_discount,
            'collection_percent': collection_pct,
            'paid_count': sum(1 for r in rows if r['status'] == 'paid'),
            'partial_count': sum(1 for r in rows if r['status'] == 'partial'),
            'unpaid_count': sum(1 for r in rows if r['status'] == 'unpaid'),
            'fully_paid_count': sum(1 for r in rows if r.get('fully_paid')),
            'students_with_dues': sum(1 for r in rows if r['due'] > 0),
            'total_overdue_amount': round(sum(r.get('overdue_amount', 0) for r in rows), 2),
            'total_due_till_date': round(sum(r.get('due_till_date', 0) for r in rows), 2),
            'defaulter_count': sum(1 for r in rows if r['behavior_tag'] == 'defaulter'),
            'late_count': sum(1 for r in rows if r['behavior_tag'] == 'late'),
            'regular_count': sum(1 for r in rows if r['behavior_tag'] == 'regular'),
            'upcoming_count': sum(1 for r in rows if r['upcoming_due_date']),
        }
    }


# ------- Fee-status export endpoints (PDF / XLSX / CSV) -------

def _fee_status_export_columns() -> List[str]:
    return ['Admission No', 'Student', 'Class', 'Section', 'Guardian', 'Phone',
            'Expected (Rs.)', 'Discount (Rs.)', 'Paid (Rs.)',
            'Due Till Date (Rs.)', 'Overdue Months', 'Total Due (Rs.)',
            'Due Date', 'Last Payment', 'Overdue Days', 'Status', 'Behavior']


def _fee_status_row_to_list(r: Dict[str, Any]) -> List[Any]:
    return [
        r.get('admission_number') or '',
        r.get('full_name') or '',
        r.get('class_name') or '',
        r.get('section') or '',
        r.get('father_name') or '',
        r.get('phone') or '',
        f"{r.get('expected', 0):,.2f}",
        f"{r.get('discount', 0):,.2f}",
        f"{r.get('paid', 0):,.2f}",
        f"{r.get('due_till_date', 0):,.2f}",
        r.get('overdue_months') or 0,
        f"{r.get('due', 0):,.2f}",
        r.get('due_date') or '',
        r.get('last_payment_date') or '',
        r.get('overdue_days') or 0,
        (r.get('status') or '').title(),
        (r.get('behavior_tag') or 'na').title(),
    ]


@api.get('/reports/monthly-dues.xlsx',
         dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant', 'owner'))])
async def report_monthly_dues_xlsx(request: Request,
                                   current=Depends(get_current_user),
                                   school_id: Optional[str] = None,
                                   class_sections: Optional[str] = None,
                                   only_with_dues: bool = True,
                                   as_of_date: Optional[str] = None,
                                   from_date: Optional[str] = None,
                                   to_date: Optional[str] = None):
    """Export a **Due List** — for each student in the selected classes, list
    monthly-fee status (paid / partial / overdue / upcoming) across all 12
    months of the academic session, along with total due amount and
    student/guardian contact details. Ideal for classroom-level follow-ups.

    Query params:
      - `class_sections`: comma-separated `class_id:section` pairs (blank
        section = all sections). Omit to export all classes.
      - `only_with_dues` (default True): if True, only students who currently
        have Due > 0 are exported.
      - `as_of_date` (YYYY-MM-DD, optional): use this as the "today" reference
        when marking months overdue / upcoming. Defaults to today.
      - `from_date` / `to_date` (YYYY-MM-DD, optional): restrict the exported
        month columns to those whose calendar (year, month) fall within
        [from_date, to_date]. Total Due is recomputed as the sum of dues from
        just those months.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = await report_fee_status(request, current, school_id, None, None,
                                   class_sections, None, None, None, None,
                                   None, None, None, as_of_date)
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    school = await schools_col.find_one({'id': sid}, {'_id': 0}) or {}

    # ---- Month range filter ----------------------------------------------
    def _yyyymm(iso: str):
        try:
            dt = datetime.fromisoformat(iso)
            return (dt.year, dt.month)
        except Exception:
            return None
    from_yyyymm = _yyyymm(from_date) if from_date else None
    to_yyyymm = _yyyymm(to_date) if to_date else None

    def _month_in_range(m):
        # `label` looks like "April 2026" — parse using calendar
        lbl = (m.get('label') or '').strip()
        try:
            dt = datetime.strptime(lbl, '%B %Y')
            key = (dt.year, dt.month)
        except Exception:
            return True  # fail-open: include if parsing fails
        if from_yyyymm and key < from_yyyymm:
            return False
        if to_yyyymm and key > to_yyyymm:
            return False
        return True

    rows = data['rows']
    # Apply month range filter and recompute per-row 'due' for those months
    range_applied = bool(from_yyyymm or to_yyyymm)
    if range_applied:
        for r in rows:
            monthly = [m for m in (r.get('monthly_status') or []) if _month_in_range(m)]
            r['monthly_status'] = monthly
            r['due'] = round(sum(max(m.get('due') or 0, 0) for m in monthly if m.get('status') in ('overdue', 'partial', 'pending')), 2)

    if only_with_dues:
        rows = [r for r in rows if (r.get('due') or 0) > 0]
    # Sort by class → section → name for easy classroom distribution
    rows.sort(key=lambda r: (r.get('class_name') or '', r.get('section') or '',
                             r.get('full_name') or ''))

    # Month headers (short form: Apr, May, ...)
    month_short = []
    if rows:
        for m in rows[0].get('monthly_status') or []:
            lbl = m.get('label') or ''
            month_short.append(lbl.split(' ')[0][:3] if lbl else '')
    if not month_short:
        month_short = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep',
                       'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Due List'

    # Title & meta
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8 + len(month_short))
    t = ws.cell(row=1, column=1, value=f"{school.get('name', 'Stanvard School')} — Fee Due List")
    t.font = Font(bold=True, size=14, color='0B2F4A')
    t.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8 + len(month_short))
    meta_parts = [
        f"Session {data.get('academic_session', '')}",
        f"{len(rows)} student(s)",
        f"Total due: Rs. {sum(r.get('due', 0) for r in rows):,.2f}",
    ]
    if as_of_date:
        meta_parts.append(f"As-of: {as_of_date}")
    if from_date or to_date:
        meta_parts.append(f"Range: {from_date or 'start'} → {to_date or 'end'}")
    meta_parts.append(f"Generated on {datetime.now().strftime('%d %b %Y, %H:%M')}")
    m = ws.cell(row=2, column=1, value=' · '.join(meta_parts))
    m.font = Font(size=10, color='475569')
    m.alignment = Alignment(horizontal='center')

    header = ['Admission No', 'Student Name', 'Class', 'Section',
              'Guardian', 'Contact', 'Monthly (Rs.)'] + month_short + ['Total Due (Rs.)']
    ws.append([])  # blank row
    ws.append(header)
    header_row_idx = ws.max_row
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='0B2F4A')
    for col_idx in range(1, len(header) + 1):
        c = ws.cell(row=header_row_idx, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Color palette
    fills = {
        'paid_full':   PatternFill('solid', fgColor='0F766E'),   # deep green — fully paid
        'paid':        PatternFill('solid', fgColor='A7F3D0'),   # light green — that month paid
        'partial':     PatternFill('solid', fgColor='FDE68A'),   # amber
        'overdue':     PatternFill('solid', fgColor='FCA5A5'),   # red
        'pending':     PatternFill('solid', fgColor='F1F5F9'),   # grey
    }
    thin = Side(border_style='thin', color='CBD5E1')
    cell_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for r in rows:
        ws.append([
            r.get('admission_number') or '',
            r.get('full_name') or '',
            r.get('class_name') or '',
            r.get('section') or '',
            r.get('father_name') or '',
            r.get('phone') or '',
            round(r.get('monthly_amount') or 0, 2),
        ] + [
            # Cell text: '✓' for paid, amount for partial/overdue, blank for pending
            (
                '✓' if m['status'] == 'paid' else
                round(m['due'], 0) if m['status'] in ('overdue', 'partial') else
                ''
            )
            for m in (r.get('monthly_status') or [])
        ] + [round(r.get('due') or 0, 2)])

        row_idx = ws.max_row
        # Colour the month cells
        for j, m in enumerate((r.get('monthly_status') or [])):
            cell = ws.cell(row=row_idx, column=8 + j)  # 7 leading cols + 1 (1-indexed)
            key = 'paid_full' if r.get('fully_paid') and m['status'] == 'paid' else m['status']
            cell.fill = fills.get(key, fills['pending'])
            cell.alignment = Alignment(horizontal='center')
            if key == 'paid_full':
                cell.font = Font(bold=True, color='FFFFFF')
            elif m['status'] == 'overdue':
                cell.font = Font(bold=True, color='B42318')
        # Border for whole row
        for col_idx in range(1, len(header) + 1):
            ws.cell(row=row_idx, column=col_idx).border = cell_border
        # Bold the last "Total Due" cell
        last = ws.cell(row=row_idx, column=len(header))
        last.font = Font(bold=True, color='B42318')
        last.alignment = Alignment(horizontal='right')

    # Column widths
    widths = [14, 30, 12, 8, 22, 14, 12] + [7] * len(month_short) + [14]
    from openpyxl.utils import get_column_letter
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.row_dimensions[header_row_idx].height = 26
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=3)

    # Legend on a second sheet
    leg = wb.create_sheet('Legend')
    leg.append(['Colour', 'Meaning'])
    for c in leg[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
    legend_rows = [
        ('Fully paid (all 12 months)', 'paid_full'),
        ('Monthly fee paid', 'paid'),
        ('Partial payment', 'partial'),
        ('Overdue (past due date)', 'overdue'),
        ('Upcoming month (not yet due)', 'pending'),
    ]
    for label, key in legend_rows:
        leg.append([' ', label])
        cell = leg.cell(row=leg.max_row, column=1)
        cell.fill = fills[key]
        cell.border = cell_border
    leg.column_dimensions['A'].width = 12
    leg.column_dimensions['B'].width = 36

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = 'due_list.xlsx'
    return StreamingResponse(
        buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


# Shared query args for all fee-status export endpoints
_FS_QP = dict(
    school_id=None, class_id=None, section=None, class_sections=None,
    status_filter=None, due_min=None, due_max=None,
    payment_date_start=None, payment_date_end=None,
    quick_view=None, behavior=None,
)


@api.get('/reports/fee-status.pdf', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant', 'owner'))])
async def report_fee_status_pdf(request: Request, current=Depends(get_current_user),
                                school_id: Optional[str] = None,
                                class_id: Optional[str] = None,
                                section: Optional[str] = None,
                                class_sections: Optional[str] = None,
                                status_filter: Optional[str] = None,
                                due_min: Optional[float] = None,
                                due_max: Optional[float] = None,
                                payment_date_start: Optional[str] = None,
                                payment_date_end: Optional[str] = None,
                                quick_view: Optional[str] = None,
                                behavior: Optional[str] = None):
    data = await report_fee_status(request, current, school_id, class_id, section, class_sections,
                                   status_filter, due_min, due_max, payment_date_start, payment_date_end,
                                   quick_view, behavior)
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    school = await schools_col.find_one({'id': sid}, {'_id': 0}) or {}
    rows = [_fee_status_row_to_list(r) for r in data['rows']]
    subtitle_bits: List[str] = []
    if class_sections:
        subtitle_bits.append(f"Classes/Sections: {class_sections}")
    if status_filter and status_filter != 'all':
        subtitle_bits.append(f"Status: {status_filter}")
    if quick_view:
        subtitle_bits.append(f"View: {quick_view}")
    if behavior and behavior != 'all':
        subtitle_bits.append(f"Behavior: {behavior}")
    subtitle = ' | '.join(subtitle_bits) if subtitle_bits else 'All Classes / Sections'
    summary = {
        'Students': data['count'],
        'Total Expected': f"Rs. {data['summary']['total_expected']:,.2f}",
        'Total Paid': f"Rs. {data['summary']['total_paid']:,.2f}",
        'Total Due': f"Rs. {data['summary']['total_due']:,.2f}",
        'Collection %': f"{data['summary']['collection_percent']}%",
        'Paid / Partial / Unpaid': (
            f"{data['summary']['paid_count']} / {data['summary']['partial_count']} / {data['summary']['unpaid_count']}"
        ),
        'Defaulters': data['summary']['defaulter_count'],
    }
    pdf = generate_report_pdf('Student Fee Status Report', subtitle,
                              _fee_status_export_columns(), rows,
                              school.get('name', 'Stanvard School'), summary)
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf',
                             headers={'Content-Disposition': 'inline; filename="fee_status_report.pdf"'})


@api.get('/reports/fee-status.xlsx', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant', 'owner'))])
async def report_fee_status_xlsx(request: Request, current=Depends(get_current_user),
                                 school_id: Optional[str] = None,
                                 class_id: Optional[str] = None,
                                 section: Optional[str] = None,
                                 class_sections: Optional[str] = None,
                                 status_filter: Optional[str] = None,
                                 due_min: Optional[float] = None,
                                 due_max: Optional[float] = None,
                                 payment_date_start: Optional[str] = None,
                                 payment_date_end: Optional[str] = None,
                                 quick_view: Optional[str] = None,
                                 behavior: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    data = await report_fee_status(request, current, school_id, class_id, section, class_sections,
                                   status_filter, due_min, due_max, payment_date_start, payment_date_end,
                                   quick_view, behavior)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Fee Status'
    header = _fee_status_export_columns()
    ws.append(header)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0B2F4A')
    for col_idx, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for r in data['rows']:
        ws.append([
            r.get('admission_number') or '',
            r.get('full_name') or '',
            r.get('class_name') or '',
            r.get('section') or '',
            r.get('father_name') or '',
            r.get('phone') or '',
            r.get('expected', 0),
            r.get('discount', 0),
            r.get('paid', 0),
            r.get('due_till_date', 0),
            r.get('overdue_months', 0),
            r.get('due', 0),
            r.get('due_date') or '',
            r.get('last_payment_date') or '',
            r.get('overdue_days') or 0,
            (r.get('status') or '').title(),
            (r.get('behavior_tag') or 'na').title(),
        ])
    widths = [16, 26, 14, 10, 22, 14, 14, 14, 14, 16, 12, 14, 12, 14, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        # Use openpyxl utils to handle >26 columns safely
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w
    s = data['summary']
    ws2 = wb.create_sheet('Summary')
    ws2.append(['Metric', 'Value'])
    for k, v in {
        'Total Students': data['count'],
        'Total Expected (Rs.)': s['total_expected'],
        'Total Paid (Rs.)': s['total_paid'],
        'Total Due (Rs.)': s['total_due'],
        'Total Discount (Rs.)': s['total_discount'],
        'Collection %': s['collection_percent'],
        'Paid Students': s['paid_count'],
        'Partial Students': s['partial_count'],
        'Unpaid Students': s['unpaid_count'],
        'Defaulters': s['defaulter_count'],
        'Late Payers': s['late_count'],
        'Regular Payers': s['regular_count'],
    }.items():
        ws2.append([k, v])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 20

    # Class rollup sheet
    ws3 = wb.create_sheet('By Class')
    ws3.append(['Class', 'Section', 'Students', 'Expected', 'Paid', 'Due', 'Collection %',
                'Paid', 'Partial', 'Unpaid', 'With Dues'])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for b in data['by_class']:
        ws3.append([b['class_name'], b['section'], b['students'],
                    b['expected'], b['paid'], b['due'], b['collection_percent'],
                    b['paid_count'], b['partial_count'], b['unpaid_count'],
                    b.get('students_with_dues', 0)])
    for col, w in enumerate([14, 10, 10, 14, 14, 14, 12, 8, 8, 8, 10], 1):
        from openpyxl.utils import get_column_letter
        ws3.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': 'attachment; filename="fee_status_report.xlsx"'})


@api.get('/reports/fee-status.csv', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant', 'owner'))])
async def report_fee_status_csv(request: Request, current=Depends(get_current_user),
                                school_id: Optional[str] = None,
                                class_id: Optional[str] = None,
                                section: Optional[str] = None,
                                class_sections: Optional[str] = None,
                                status_filter: Optional[str] = None,
                                due_min: Optional[float] = None,
                                due_max: Optional[float] = None,
                                payment_date_start: Optional[str] = None,
                                payment_date_end: Optional[str] = None,
                                quick_view: Optional[str] = None,
                                behavior: Optional[str] = None):
    import csv
    data = await report_fee_status(request, current, school_id, class_id, section, class_sections,
                                   status_filter, due_min, due_max, payment_date_start, payment_date_end,
                                   quick_view, behavior)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_fee_status_export_columns())
    for r in data['rows']:
        writer.writerow(_fee_status_row_to_list(r))
    return StreamingResponse(io.BytesIO(buffer.getvalue().encode()),
                             media_type='text/csv',
                             headers={'Content-Disposition': 'attachment; filename="fee_status_report.csv"'})


# =====================================================
# FEE ANALYTICS (dedicated, filterable)
# =====================================================
@api.get('/analytics/fees')
async def fee_analytics(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        start_date: Optional[str] = None,
                        end_date: Optional[str] = None,
                        class_id: Optional[str] = None,
                        section: Optional[str] = None,
                        payment_mode: Optional[str] = None,
                        payment_status: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    today_iso = datetime.now().strftime('%Y-%m-%d')

    # Payment query
    pq: Dict[str, Any] = {'school_id': sid, 'status': 'success'}
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date + 'T23:59:59'
        pq['paid_at'] = rq
    if payment_mode:
        pq['payment_mode'] = payment_mode
    payments_all = await payments_col.find(pq, {'_id': 0}).to_list(50000)

    # Filter by class/section (need student data)
    students_all = await students_col.find({'school_id': sid}, {'_id': 0}).to_list(20000)
    stu_by_id = {s['id']: s for s in students_all}
    def _in_scope(pmt):
        s = stu_by_id.get(pmt.get('student_id'))
        if not s:
            return False
        if class_id and s.get('class_id') != class_id:
            return False
        if section and s.get('section') != section:
            return False
        return True
    payments = [p for p in payments_all if _in_scope(p)]

    total_collected = sum(p.get('total_paid', 0) for p in payments)
    total_discount = sum(p.get('discount', 0) for p in payments)
    total_late_fee = sum(p.get('late_fee', 0) for p in payments)

    # Today / month specific (independent of the filter for context KPIs)
    month_start = datetime.now().strftime('%Y-%m-01')
    today_payments = await payments_col.find({'school_id': sid, 'status': 'success',
                                              'paid_at': {'$gte': today_iso, '$lte': today_iso + 'T23:59:59'}}, {'_id': 0}).to_list(5000)
    month_payments = await payments_col.find({'school_id': sid, 'status': 'success',
                                              'paid_at': {'$gte': month_start}}, {'_id': 0}).to_list(20000)

    # Expected & pending computation (fee-status logic, filtered by class/section)
    students_scope = [s for s in students_all if s.get('status') == 'active'
                      and (not class_id or s.get('class_id') == class_id)
                      and (not section or s.get('section') == section)]
    student_ids_scope = [s['id'] for s in students_scope]
    assignments = await fee_assignments_col.find({'student_id': {'$in': student_ids_scope}}, {'_id': 0}).to_list(50000)
    plan_cache: Dict[str, Any] = {}
    async def _get_plan(pid):
        if pid not in plan_cache:
            plan_cache[pid] = await fee_plans_col.find_one({'id': pid}, {'_id': 0}) or {}
        return plan_cache[pid]

    expected_by_stu: Dict[str, float] = {}
    discount_by_stu: Dict[str, float] = {}
    for a in assignments:
        items = a.get('custom_items') or []
        if not items and a.get('fee_plan_id'):
            plan = await _get_plan(a['fee_plan_id'])
            items = plan.get('items', [])
        expected_by_stu[a['student_id']] = expected_by_stu.get(a['student_id'], 0.0) + sum(i.get('amount', 0) for i in items)
        discount_by_stu[a['student_id']] = discount_by_stu.get(a['student_id'], 0.0) + a.get('discount_amount', 0)

    paid_by_stu: Dict[str, float] = {}
    all_paid_by_stu = await payments_col.find({'student_id': {'$in': student_ids_scope}, 'status': 'success'}, {'_id': 0}).to_list(50000)
    for p in all_paid_by_stu:
        paid_by_stu[p['student_id']] = paid_by_stu.get(p['student_id'], 0.0) + p.get('total_paid', 0)

    paid_count = 0
    partial_count = 0
    unpaid_count = 0
    total_expected = 0.0
    total_pending = 0.0
    for s in students_scope:
        exp = expected_by_stu.get(s['id'], 0.0)
        disc = discount_by_stu.get(s['id'], 0.0)
        pd = paid_by_stu.get(s['id'], 0.0)
        due = max(exp - disc - pd, 0.0)
        total_expected += exp
        total_pending += due
        if pd <= 0:
            unpaid_count += 1
        elif due <= 0:
            paid_count += 1
        else:
            partial_count += 1

    # Apply payment_status filter to KPIs if requested (only for pending/paid student counts)
    # (already computed all; payment_status just narrows which student cards are displayed - handled client-side)

    # Daily buckets across the filter range (or last 14 days if no range)
    if start_date and end_date:
        try:
            d0 = datetime.strptime(start_date, '%Y-%m-%d')
            d1 = datetime.strptime(end_date, '%Y-%m-%d')
        except Exception:
            d0 = datetime.now() - timedelta(days=13)
            d1 = datetime.now()
    else:
        d0 = datetime.now() - timedelta(days=13)
        d1 = datetime.now()
    days = min((d1 - d0).days + 1, 92)  # cap at ~3 months for the daily view
    daily = []
    for i in range(days):
        d = (d0 + timedelta(days=i)).strftime('%Y-%m-%d')
        amt = sum(p.get('total_paid', 0) for p in payments if (p.get('paid_at') or '').startswith(d))
        cnt = sum(1 for p in payments if (p.get('paid_at') or '').startswith(d))
        daily.append({'date': d, 'amount': amt, 'transactions': cnt})

    # Monthly (12 months of current year - independent of filter)
    year = datetime.now().year
    monthly = []
    for m in range(1, 13):
        m_start = f'{year}-{m:02d}-01'
        m_end = f'{year}-{m + 1:02d}-01' if m < 12 else f'{year + 1}-01-01'
        mpays = [p for p in payments_all if m_start <= (p.get('paid_at') or '') < m_end]
        monthly.append({
            'month': datetime(year, m, 1).strftime('%b'),
            'amount': sum(p.get('total_paid', 0) for p in mpays),
            'transactions': len(mpays),
        })

    # Payment mode breakdown
    by_mode: Dict[str, Dict[str, Any]] = {}
    for p in payments:
        m = p.get('payment_mode', 'cash')
        b = by_mode.setdefault(m, {'amount': 0.0, 'count': 0})
        b['amount'] += p.get('total_paid', 0)
        b['count'] += 1

    # Class-wise collection
    classes = await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(200)
    cls_map = {c['id']: c['name'] for c in classes}
    by_class_map: Dict[str, Dict[str, Any]] = {}
    for p in payments:
        s = stu_by_id.get(p.get('student_id'))
        if not s:
            continue
        cname = cls_map.get(s.get('class_id'), '—')
        b = by_class_map.setdefault(cname, {'amount': 0.0, 'transactions': 0, 'students': set()})
        b['amount'] += p.get('total_paid', 0)
        b['transactions'] += 1
        b['students'].add(s['id'])
    by_class = sorted(
        [{'class_name': k, 'amount': v['amount'], 'transactions': v['transactions'],
          'students': len(v['students'])} for k, v in by_class_map.items()],
        key=lambda x: x['amount'], reverse=True,
    )

    # Transactions list (enriched, sorted by paid_at desc) for the currently applied filter
    transactions = []
    for p in payments:
        s = stu_by_id.get(p.get('student_id')) or {}
        items = p.get('items') or []
        # Compose a compact fee-heads label (e.g., "Tuition, Transport")
        fee_heads = ', '.join(
            [i.get('fee_head_name') or '' for i in items if i.get('fee_head_name')]
        ) or '—'
        transactions.append({
            'id': p.get('id'),
            'receipt_number': p.get('receipt_number') or '—',
            'paid_at': p.get('paid_at'),
            'student_id': p.get('student_id'),
            'student_name': p.get('student_name') or s.get('full_name') or '—',
            'admission_number': s.get('admission_number') or '—',
            'class_name': cls_map.get(s.get('class_id'), '—'),
            'section': s.get('section') or '—',
            'father_name': s.get('father_name') or '—',
            'phone': s.get('phone') or '—',
            'fee_heads': fee_heads,
            'subtotal': p.get('subtotal', 0),
            'discount': p.get('discount', 0),
            'late_fee': p.get('late_fee', 0),
            'total_paid': p.get('total_paid', 0),
            'payment_mode': p.get('payment_mode', 'cash'),
            'txn_ref': p.get('txn_ref') or '',
            'status': p.get('status', 'success'),
            'collected_by_name': p.get('collected_by_name') or '—',
            'remarks': p.get('remarks') or '',
        })
    # Sort newest first
    transactions.sort(key=lambda t: (t.get('paid_at') or ''), reverse=True)

    return {
        'kpis': {
            'total_collected': total_collected,
            'total_pending': total_pending,
            'total_expected': total_expected,
            'total_paid_students': paid_count,
            'total_partial_students': partial_count,
            'total_pending_students': unpaid_count,
            'today_collection': sum(p.get('total_paid', 0) for p in today_payments),
            'today_transactions': len(today_payments),
            'monthly_collection': sum(p.get('total_paid', 0) for p in month_payments),
            'monthly_transactions': len(month_payments),
            'total_discount': total_discount,
            'total_late_fee': total_late_fee,
            'transactions_in_range': len(payments),
        },
        'daily': daily,
        'monthly': monthly,
        'by_mode': by_mode,
        'by_class': by_class,
        'transactions': transactions,
        'range': {'start_date': start_date, 'end_date': end_date},
    }


@api.get('/analytics/student/{student_id}/fee-report')
async def student_fee_report(student_id: str, request: Request, current=Depends(get_current_user)):
    """Full fee profile for a single student."""
    student = await students_col.find_one({'id': student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    if current['role'] != 'super_admin' and student['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')

    assignments = await fee_assignments_col.find({'student_id': student_id}, {'_id': 0}).to_list(50)
    plan_cache: Dict[str, Any] = {}
    total_expected = 0.0
    total_discount = 0.0
    due_dates: List[str] = []
    line_items: List[Dict[str, Any]] = []
    for a in assignments:
        items = a.get('custom_items') or []
        if not items and a.get('fee_plan_id'):
            if a['fee_plan_id'] not in plan_cache:
                plan_cache[a['fee_plan_id']] = await fee_plans_col.find_one({'id': a['fee_plan_id']}, {'_id': 0}) or {}
            items = plan_cache[a['fee_plan_id']].get('items', [])
        for it in items:
            line_items.append({
                'fee_head_name': it.get('fee_head_name'),
                'frequency': it.get('frequency', 'monthly'),
                'amount': it.get('amount', 0),
                'due_date': it.get('due_date') or a.get('due_date'),
            })
            total_expected += it.get('amount', 0)
        total_discount += a.get('discount_amount', 0)
        if a.get('due_date'):
            due_dates.append(a['due_date'])

    payments = await payments_col.find({'student_id': student_id, 'status': 'success'}, {'_id': 0}).sort('paid_at', -1).to_list(500)
    total_paid = sum(p.get('total_paid', 0) for p in payments)
    total_late_paid = sum(p.get('late_fee', 0) for p in payments)
    balance = max(total_expected - total_discount - total_paid, 0)

    today = datetime.now().date()
    next_due_date = None
    days_overdue = 0
    if due_dates:
        due_sorted = sorted(due_dates)
        next_due_date = due_sorted[0]
        try:
            dd = datetime.strptime(next_due_date, '%Y-%m-%d').date()
            if dd < today and balance > 0:
                days_overdue = (today - dd).days
        except Exception:
            pass

    status = 'unpaid' if total_paid == 0 else ('paid' if balance <= 0 else 'partial')
    last_payment_date = payments[0]['paid_at'][:10] if payments else None

    # Class name
    class_name = '—'
    if student.get('class_id'):
        c = await classes_col.find_one({'id': student['class_id']}, {'_id': 0})
        if c:
            class_name = c['name']

    return {
        'student': {**student, 'class_name': class_name},
        'summary': {
            'total_expected': total_expected,
            'total_discount': total_discount,
            'total_paid': total_paid,
            'total_late_paid': total_late_paid,
            'balance': balance,
            'status': status,
            'last_payment_date': last_payment_date,
            'next_due_date': next_due_date,
            'days_overdue': days_overdue,
        },
        'line_items': line_items,
        'payments': payments,
    }


@api.get('/analytics/student/{student_id}/fee-report.pdf')
async def student_fee_report_pdf(student_id: str, request: Request, current=Depends(get_current_user)):
    data = await student_fee_report(student_id, request, current)
    school = await schools_col.find_one({'id': data['student']['school_id']}, {'_id': 0}) or {}
    s = data['student']
    summary = data['summary']

    cols = ['Receipt No', 'Date', 'Mode', 'Late Fee', 'Discount', 'Amount (Rs.)']
    rows = []
    for p in data['payments']:
        rows.append([
            p.get('receipt_number', ''),
            (p.get('paid_at') or '')[:10],
            str(p.get('payment_mode', '')).replace('_', ' ').title(),
            f"{p.get('late_fee', 0):,.2f}",
            f"{p.get('discount', 0):,.2f}",
            f"{p.get('total_paid', 0):,.2f}",
        ])
    subtitle = (f"Student Fee Report — {s.get('full_name')} ({s.get('admission_number')}) — "
               f"Class {s.get('class_name')} {s.get('section') or ''}")
    summary_line = {
        'Total Fee': f"Rs. {summary['total_expected']:,.2f}",
        'Total Paid': f"Rs. {summary['total_paid']:,.2f}",
        'Discount': f"Rs. {summary['total_discount']:,.2f}",
        'Late Fee': f"Rs. {summary['total_late_paid']:,.2f}",
        'Balance': f"Rs. {summary['balance']:,.2f}",
        'Status': summary['status'].upper(),
    }
    if summary.get('next_due_date'):
        summary_line['Next Due'] = summary['next_due_date']
    pdf = generate_report_pdf('Student Fee Report', subtitle, cols, rows,
                              school.get('name', 'Stanvard School'), summary_line)
    fname = f'fee_report_{s.get("admission_number", student_id)}.pdf'
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf',
                             headers={'Content-Disposition': f'inline; filename="{fname}"'})


# =====================================================
# EXAMINATIONS & REPORT CARDS
# =====================================================
def _grade_for(pct: float) -> str:
    if pct >= 90:
        return 'A+'
    if pct >= 75:
        return 'A'
    if pct >= 60:
        return 'B'
    if pct >= 45:
        return 'C'
    if pct >= 33:
        return 'D'
    return 'F'


@api.get('/exams')
async def list_exams(request: Request, current=Depends(get_current_user),
                     class_id: Optional[str] = None,
                     school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if class_id:
        q['class_id'] = class_id
    return await exams_col.find(q, {'_id': 0}).sort('created_at', -1).to_list(200)


@api.post('/exams', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_exam(body: ExamCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    subjects = [s for s in body.subjects if s.name and s.name.strip()]
    if not subjects:
        raise HTTPException(400, 'At least one subject is required')
    lowered = [s.name.strip().lower() for s in subjects]
    if len(lowered) != len(set(lowered)):
        raise HTTPException(400, 'Subject names must be unique')
    ex = Exam(school_id=sid, name=body.name.strip(), term=body.term or '',
              academic_session=body.academic_session, class_id=body.class_id,
              section=body.section or None, subjects=subjects, exam_date=body.exam_date)
    await exams_col.insert_one(ex.model_dump())
    await log_audit(action='exam.create', current_user=current, school_id=sid,
                    entity_type='exam', entity_id=ex.id, details={'name': ex.name})
    return ex.model_dump()


@api.get('/exams/{exam_id}')
async def get_exam(exam_id: str, current=Depends(get_current_user)):
    exam = await exams_col.find_one({'id': exam_id}, {'_id': 0})
    if not exam:
        raise HTTPException(404, 'Exam not found')
    if current['role'] != 'super_admin' and exam['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    q: Dict[str, Any] = {'school_id': exam['school_id'], 'class_id': exam['class_id'],
                         'status': 'active'}
    if exam.get('section'):
        q['section'] = exam['section']
    students = await students_col.find(q, {'_id': 0}).to_list(2000)
    students.sort(key=lambda s: (s.get('section') or '', str(s.get('roll_number') or ''),
                                 s.get('full_name') or ''))
    marks_docs = await marks_col.find({'exam_id': exam_id}, {'_id': 0}).to_list(5000)
    return {'exam': exam, 'students': students,
            'marks': {m['student_id']: m for m in marks_docs}}


@api.patch('/exams/{exam_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_exam(exam_id: str, body: ExamUpdate, current=Depends(get_current_user)):
    upd = body.model_dump(exclude_none=True)
    if not upd:
        return await exams_col.find_one({'id': exam_id}, {'_id': 0})
    upd['updated_at'] = now_iso()
    r = await exams_col.update_one({'id': exam_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'Exam not found')
    await log_audit(action='exam.update', current_user=current,
                    entity_type='exam', entity_id=exam_id, details=upd)
    return await exams_col.find_one({'id': exam_id}, {'_id': 0})


@api.delete('/exams/{exam_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_exam(exam_id: str, current=Depends(get_current_user)):
    r = await exams_col.delete_one({'id': exam_id})
    if not r.deleted_count:
        raise HTTPException(404, 'Exam not found')
    await marks_col.delete_many({'exam_id': exam_id})
    await log_audit(action='exam.delete', current_user=current,
                    entity_type='exam', entity_id=exam_id)
    return {'ok': True}


@api.post('/exams/{exam_id}/marks',
          dependencies=[Depends(require_roles('super_admin', 'school_admin', 'teacher'))])
async def save_exam_marks(exam_id: str, body: MarksBulkSave, current=Depends(get_current_user)):
    exam = await exams_col.find_one({'id': exam_id}, {'_id': 0})
    if not exam:
        raise HTTPException(404, 'Exam not found')
    if current['role'] != 'super_admin' and exam['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    subj_max = {s['name']: float(s.get('max_marks') or 100) for s in (exam.get('subjects') or [])}
    if not subj_max:
        raise HTTPException(400, 'Exam has no subjects configured')
    saved = 0
    for entry in body.entries:
        marks: Dict[str, Any] = {}
        total = 0.0
        max_total = 0.0
        for name, mx in subj_max.items():
            v = entry.marks.get(name)
            max_total += mx
            if v is None:
                marks[name] = None
                continue
            v = float(v)
            if not (0 <= v <= mx):
                raise HTTPException(400, f'Marks for "{name}" must be between 0 and {mx:g}')
            marks[name] = round(v, 2)
            total += v
        pct = round(total / max_total * 100, 2) if max_total else 0.0
        await marks_col.update_one(
            {'exam_id': exam_id, 'student_id': entry.student_id},
            {'$set': {
                'school_id': exam['school_id'], 'exam_id': exam_id,
                'student_id': entry.student_id, 'marks': marks,
                'total': round(total, 2), 'max_total': round(max_total, 2),
                'percentage': pct, 'grade': _grade_for(pct),
                'entered_by_id': current.get('id'),
                'entered_by_name': current.get('full_name'),
                'updated_at': now_iso(),
            }, '$setOnInsert': {'id': new_id(), 'created_at': now_iso()}},
            upsert=True,
        )
        saved += 1
    await log_audit(action='exam.marks_save', current_user=current, school_id=exam['school_id'],
                    entity_type='exam', entity_id=exam_id, details={'entries': saved})
    return {'ok': True, 'saved': saved}


@api.get('/exams/{exam_id}/results')
async def exam_results(exam_id: str, current=Depends(get_current_user)):
    data = await get_exam(exam_id, current)
    marks_map = data['marks']
    results = []
    for s in data['students']:
        m = marks_map.get(s['id']) or {}
        results.append({
            'student_id': s['id'], 'full_name': s.get('full_name'),
            'admission_number': s.get('admission_number'),
            'roll_number': s.get('roll_number'), 'section': s.get('section'),
            'marks': m.get('marks') or {},
            'total': m.get('total', 0), 'max_total': m.get('max_total', 0),
            'percentage': m.get('percentage', 0), 'grade': m.get('grade'),
            'has_marks': bool(m),
        })
    ranked = sorted((r for r in results if r['has_marks']),
                    key=lambda r: r['total'], reverse=True)
    for i, r in enumerate(ranked):
        r['rank'] = i + 1
    return {'exam': data['exam'], 'results': results}


@api.get('/students/{student_id}/report-card')
async def student_report_card(student_id: str, current=Depends(get_current_user)):
    student = await students_col.find_one({'id': student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    if current['role'] != 'super_admin' and student['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')
    exams = await exams_col.find(
        {'school_id': student['school_id'], 'class_id': student.get('class_id')},
        {'_id': 0}).sort('exam_date', 1).to_list(200)
    out = []
    for ex in exams:
        if ex.get('section') and student.get('section') and ex['section'] != student['section']:
            continue
        m = await marks_col.find_one({'exam_id': ex['id'], 'student_id': student_id}, {'_id': 0})
        rows = []
        for subj in (ex.get('subjects') or []):
            mx = float(subj.get('max_marks') or 100)
            v = ((m or {}).get('marks') or {}).get(subj['name'])
            if v is None:
                grade = 'AB' if m else '-'
            else:
                grade = _grade_for(round(float(v) / mx * 100, 2) if mx else 0)
            rows.append({'subject': subj['name'], 'max_marks': mx, 'marks': v, 'grade': grade})
        out.append({'exam': ex, 'rows': rows, 'marks_entered': bool(m),
                    'total': (m or {}).get('total'), 'max_total': (m or {}).get('max_total'),
                    'percentage': (m or {}).get('percentage'), 'grade': (m or {}).get('grade')})
    return {'student': student, 'exams': out}


@api.get('/students/{student_id}/report-card.pdf')
async def student_report_card_pdf(student_id: str, exam_id: str = Query(...),
                                  current=Depends(get_current_user)):
    card = await student_report_card(student_id, current)
    entry = next((e for e in card['exams'] if e['exam']['id'] == exam_id), None)
    if not entry:
        raise HTTPException(404, 'Report card not found for this exam')
    school = await schools_col.find_one({'id': card['student']['school_id']}, {'_id': 0}) or {}
    s = card['student']
    ex = entry['exam']
    cols = ['Subject', 'Max Marks', 'Marks Obtained', 'Grade']
    rows = []
    for r in entry['rows']:
        if r['marks'] is None:
            mo = 'Absent' if entry['marks_entered'] else '-'
        else:
            mo = f"{r['marks']:g}"
        rows.append([r['subject'], f"{r['max_marks']:g}", mo, r['grade']])
    if entry['marks_entered']:
        rows.append(['TOTAL', f"{entry['max_total']:g}", f"{entry['total']:g}",
                     entry.get('grade') or ''])
    cls = await classes_col.find_one({'id': s.get('class_id')}, {'_id': 0}) or {}
    subtitle = (f"{ex['name']} ({ex.get('academic_session', '')}) - {s.get('full_name')} "
                f"({s.get('admission_number')}) - {cls.get('name', '')} {s.get('section') or ''}")
    summary = {'Percentage': f"{entry.get('percentage') or 0}%",
               'Grade': entry.get('grade') or '-'}
    pdf = generate_report_pdf('STUDENT REPORT CARD', subtitle, cols, rows,
                              school.get('name', 'Stanvard School'), summary)
    fname = f"report_card_{s.get('admission_number', student_id)}.pdf"
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf',
                             headers={'Content-Disposition': f'inline; filename="{fname}"'})


# =====================================================
# BULK STUDENT IMPORT (CSV)
# =====================================================
@api.post('/students/import', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def import_students(request: Request, file: UploadFile = File(...),
                          current=Depends(get_current_user)):
    sid = resolve_school_id(current, None, request.headers.get('X-School-Id'))
    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(400, 'File too large (max 5 MB)')
    try:
        text = raw.decode('utf-8-sig')
    except Exception:
        raise HTTPException(400, 'File must be a UTF-8 encoded CSV')
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames or 'full_name' not in [f.strip().lower() for f in reader.fieldnames]:
        raise HTTPException(400, 'CSV must contain a "full_name" column')
    classes = await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(1000)
    class_by_name = {str(c.get('name', '')).strip().lower(): c for c in classes}
    school = await schools_col.find_one({'id': sid}, {'_id': 0}) or {}
    code = school.get('code', 'STV')
    sess_year = (school.get('academic_session') or '2025-26').split('-')[0]
    created = 0
    errors: List[str] = []
    for line_no, row in enumerate(reader, start=2):
        row = {(k or '').strip().lower(): (v or '').strip() for k, v in row.items()}
        name = row.get('full_name')
        if not name:
            errors.append(f'Row {line_no}: full_name is empty - skipped')
            continue
        cls = None
        cname = row.get('class_name')
        if cname:
            cls = class_by_name.get(cname.lower())
            if not cls:
                errors.append(f'Row {line_no}: class "{cname}" not found - skipped')
                continue
        adm_no = row.get('admission_number')
        if not adm_no:
            # Auto-generate; the counter may lag behind manually seeded
            # admission numbers, so skip forward until we find a free one.
            for _ in range(2000):
                seq = await get_next_sequence(f'adm_{sid}')
                adm_no = f'{code}-{sess_year}-{seq:04d}'
                if not await students_col.find_one({'school_id': sid, 'admission_number': adm_no}, {'_id': 0, 'id': 1}):
                    break
            else:
                errors.append(f'Row {line_no}: could not allocate admission number - skipped')
                continue
        dup = await students_col.find_one({'school_id': sid, 'admission_number': adm_no},
                                          {'_id': 0, 'id': 1})
        if dup:
            errors.append(f'Row {line_no}: admission number {adm_no} already exists - skipped')
            continue
        student = Student(
            school_id=sid, admission_number=adm_no, full_name=name,
            class_id=cls['id'] if cls else None,
            section=row.get('section') or None,
            roll_number=row.get('roll_number') or None,
            dob=row.get('dob') or None,
            gender=row.get('gender') or None,
            father_name=row.get('father_name') or None,
            mother_name=row.get('mother_name') or None,
            phone=row.get('phone') or None,
            email=row.get('email') or None,
        )
        await students_col.insert_one(student.model_dump())
        created += 1
    await log_audit(action='student.import', current_user=current, school_id=sid,
                    details={'created': created, 'skipped': len(errors),
                             'filename': file.filename})
    return {'created': created, 'skipped': len(errors), 'errors': errors[:50]}


# =====================================================
# TWILIO MESSAGING — SMS + WhatsApp fee reminders
# =====================================================
TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID', '')
TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN', '')
TWILIO_PHONE_NUMBER = os.environ.get('TWILIO_PHONE_NUMBER', '')
TWILIO_WHATSAPP_FROM = os.environ.get('TWILIO_WHATSAPP_FROM') or TWILIO_PHONE_NUMBER
TWILIO_CONFIGURED = bool(TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN and TWILIO_PHONE_NUMBER)

_twilio_client = None


def twilio_client():
    global _twilio_client
    if _twilio_client is None and TWILIO_CONFIGURED:
        from twilio.rest import Client
        _twilio_client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
    return _twilio_client


def _normalize_phone(raw: Optional[str]) -> Optional[str]:
    """Normalize to E.164. Bare 10-digit numbers are treated as Indian (+91)."""
    if not raw:
        return None
    digits = re.sub(r'\D', '', raw)
    if digits.startswith('0') and len(digits) == 11:
        digits = digits[1:]
    if len(digits) == 10:
        digits = '91' + digits
    if not (11 <= len(digits) <= 15):
        return None
    return f'+{digits}'


async def _twilio_send(to_e164: str, body: str, channel: str) -> Dict[str, Any]:
    client = twilio_client()
    if not client:
        return {'ok': False, 'error': 'Twilio is not configured'}

    def _do():
        if channel == 'whatsapp':
            return client.messages.create(from_=f'whatsapp:{TWILIO_WHATSAPP_FROM}',
                                          to=f'whatsapp:{to_e164}', body=body)
        return client.messages.create(from_=TWILIO_PHONE_NUMBER, to=to_e164, body=body)

    try:
        msg = await asyncio.to_thread(_do)
        return {'ok': True, 'sid': msg.sid, 'status': msg.status}
    except Exception as e:
        return {'ok': False, 'error': str(e)[:300]}


@api.get('/messaging/status')
async def messaging_status(current=Depends(get_current_user)):
    return {'configured': TWILIO_CONFIGURED,
            'channels': ['sms', 'whatsapp'] if TWILIO_CONFIGURED else [],
            'note': ('Trial account: SMS delivers only to numbers verified in the '
                     'Twilio console; WhatsApp requires joining the sandbox.')} if TWILIO_CONFIGURED else {
                'configured': False, 'channels': [], 'note': 'Twilio is not configured'}


async def _pending_for_month(sid: str, month: int, request: Request,
                             current: Dict[str, Any],
                             class_id: Optional[str] = None) -> List[Dict[str, Any]]:
    """Students whose fee for the given calendar month is still unpaid."""
    q: Dict[str, Any] = {'school_id': sid, 'status': 'active'}
    if class_id:
        q['class_id'] = class_id
    students = await students_col.find(q, {'_id': 0}).to_list(2000)
    class_map = {c['id']: c.get('name', '')
                 for c in await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(500)}
    out: List[Dict[str, Any]] = []
    for s in students:
        try:
            sched = await student_fee_schedule(s['id'], request, current)
        except Exception:
            continue
        for m in (sched.get('schedule') or []):
            if int(m.get('month') or 0) != month:
                continue
            remaining = round(float(m.get('amount') or 0) - float(m.get('paid_amount') or 0), 2)
            if remaining > 0 and m.get('status') not in ('no_fee', 'paid'):
                out.append({
                    'student_id': s['id'], 'student_name': s.get('full_name'),
                    'admission_number': s.get('admission_number'),
                    'class_name': class_map.get(s.get('class_id'), ''),
                    'phone': s.get('phone') or '', 'remaining': remaining,
                    'label': m.get('label'), 'due_date': m.get('due_date'),
                    'status': m.get('status'),
                })
            break
    return out


@api.get('/messaging/pending')
async def messaging_pending(request: Request, current=Depends(get_current_user),
                            month: int = Query(..., ge=1, le=12),
                            class_id: Optional[str] = None,
                            school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    rows = await _pending_for_month(sid, month, request, current, class_id)
    return {'count': len(rows),
            'total_amount': round(sum(r['remaining'] for r in rows), 2),
            'rows': rows[:200]}


@api.post('/messaging/fee-reminders',
          dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant'))])
async def send_fee_reminders(body: FeeReminderRequest, request: Request,
                             current=Depends(get_current_user)):
    if not TWILIO_CONFIGURED:
        raise HTTPException(503, 'Twilio is not configured')
    if not (1 <= body.month <= 12):
        raise HTTPException(400, 'month must be 1-12')
    if body.channel not in ('sms', 'whatsapp', 'both'):
        raise HTTPException(400, 'channel must be sms, whatsapp or both')
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    school = await schools_col.find_one({'id': sid}, {'_id': 0}) or {}
    school_name = school.get('name', 'School')
    rows = await _pending_for_month(sid, body.month, request, current, body.class_id)
    if len(rows) > 500:
        # Safety cap: never fan out more than 500 Twilio calls in one request.
        rows = rows[:500]
    channels = ['sms', 'whatsapp'] if body.channel == 'both' else [body.channel]
    sent = failed = skipped = 0
    results: List[Dict[str, Any]] = []
    for r in rows:
        to = _normalize_phone(r['phone'])
        if not to:
            skipped += 1
            results.append({**r, 'sent': False, 'error': 'no valid phone number'})
            continue
        text = (f"Dear Parent, the school fee of Rs.{r['remaining']:,.0f} for "
                f"{r['student_name']} ({r['class_name']}) for {r['label']} is pending.")
        if r['due_date']:
            text += f" Kindly pay by {r['due_date']}."
        text += f" - {school_name}"
        entry: Dict[str, Any] = {**r, 'to': to, 'channels': {}}
        ok_any = False
        for ch in channels:
            res = await _twilio_send(to, text, ch)
            entry['channels'][ch] = res
            ok_any = ok_any or res['ok']
            await message_logs_col.insert_one({
                'id': new_id(), 'school_id': sid, 'student_id': r['student_id'],
                'channel': ch, 'to': to, 'body': text, 'ok': res['ok'],
                'sid': res.get('sid'), 'status': res.get('status'),
                'error': res.get('error'), 'kind': 'fee_reminder',
                'sent_by': current.get('id'), 'created_at': now_iso(),
            })
        entry['sent'] = ok_any
        if ok_any:
            sent += 1
        else:
            failed += 1
        results.append(entry)
    await log_audit(action='messaging.fee_reminders', current_user=current, school_id=sid,
                    details={'month': body.month, 'channel': body.channel,
                             'pending': len(rows), 'sent': sent,
                             'failed': failed, 'skipped': skipped})
    return {'month': body.month, 'channel': body.channel, 'pending_count': len(rows),
            'sent': sent, 'failed': failed, 'skipped': skipped, 'results': results[:100]}


@api.post('/messaging/test', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def messaging_test(current=Depends(get_current_user),
                         phone: str = Query(...), channel: str = Query('sms'),
                         body: str = Query('Test message from Stanvard School ERP')):
    if not TWILIO_CONFIGURED:
        raise HTTPException(503, 'Twilio is not configured')
    to = _normalize_phone(phone)
    if not to:
        raise HTTPException(400, 'Invalid phone number')
    res = await _twilio_send(to, body, 'whatsapp' if channel == 'whatsapp' else 'sms')
    await message_logs_col.insert_one({
        'id': new_id(), 'school_id': current.get('school_id'), 'channel': channel,
        'to': to, 'body': body, 'ok': res['ok'], 'sid': res.get('sid'),
        'status': res.get('status'), 'error': res.get('error'), 'kind': 'test',
        'sent_by': current.get('id'), 'created_at': now_iso(),
    })
    if not res['ok']:
        # Return 200 with ok:false so the UI/caller gets a clean JSON error
        # (a 502 would be intercepted by the ingress and served as HTML).
        return {'ok': False, 'error': res['error'], 'to': to, 'channel': channel}
    return res


# ------------ Mount router & middlewares ------------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=['*'],
    allow_headers=['*'],
)


async def _ensure_owner_accounts():
    """Idempotent owner seeding so production gets the owner logins on deploy.
    Scoped to the primary (KNP) branch. Safe to run on every startup."""
    owners = [
        {'full_name': 'Satya Prakash Mundra', 'email': 'spmundra@stanvard.school',
         'password': 'mundra@sp2026', 'phone': '9414150001'},
        {'full_name': 'Mrityunjay Mundra', 'email': 'mjmundra@stanvard.school',
         'password': 'mundra@mj2026', 'phone': '9414150002'},
    ]
    school = await schools_col.find_one({'code': 'KNP', 'status': {'$ne': 'deleted'}}, {'_id': 0})
    if not school:
        school = await schools_col.find_one({'status': {'$ne': 'deleted'}}, {'_id': 0}, sort=[('code', 1)])
    school_id = school['id'] if school else None
    for o in owners:
        existing = await users_col.find_one({'email': o['email']}, {'_id': 0})
        if existing:
            continue  # never overwrite an existing account's password on boot
        u = User(full_name=o['full_name'], email=o['email'], phone=o['phone'],
                 role='owner', school_id=school_id,
                 password_hash=hash_password(o['password']), status='active')
        await users_col.insert_one(u.model_dump())
        logger.info('Seeded owner account: %s', o['email'])


async def _ensure_plan_status():
    """Backfill LIVE/DRAFT status on legacy fee plans that predate the field.
    For each class+session group the most-recently-created plan becomes live,
    the rest draft. Only touches plans missing a 'status' field (idempotent)."""
    legacy = await fee_plans_col.find({'status': {'$exists': False}}, {'_id': 0}).to_list(100000)
    if not legacy:
        return
    groups: Dict[tuple, list] = {}
    for p in legacy:
        key = (p.get('school_id'), p.get('class_id'), p.get('academic_session'))
        groups.setdefault(key, []).append(p)
    for arr in groups.values():
        arr.sort(key=lambda x: x.get('created_at') or '', reverse=True)
        for i, p in enumerate(arr):
            st = 'live' if (i == 0 and p.get('class_id')) else 'draft'
            await fee_plans_col.update_one({'id': p['id']},
                                           {'$set': {'status': st, 'is_active': st == 'live'}})
    logger.info('Backfilled LIVE/DRAFT status on %d legacy fee plans', len(legacy))


@app.on_event('startup')
async def startup():
    # Create indexes
    await users_col.create_index('email', unique=True)
    await students_col.create_index([('school_id', 1), ('admission_number', 1)])
    await students_col.create_index([('school_id', 1), ('class_id', 1)])
    await payments_col.create_index([('school_id', 1), ('paid_at', -1)])
    await attendance_col.create_index([('school_id', 1), ('date', 1), ('class_id', 1)])
    await exams_col.create_index([('school_id', 1), ('class_id', 1)])
    await marks_col.create_index([('exam_id', 1), ('student_id', 1)], unique=True)
    await message_logs_col.create_index([('school_id', 1), ('created_at', -1)])
    try:
        await _ensure_owner_accounts()
    except Exception as e:
        logger.exception('owner seeding skipped: %s', e)
    try:
        await _ensure_plan_status()
    except Exception as e:
        logger.exception('plan-status backfill skipped: %s', e)
    logger.info('Stanvard ERP API started')


@app.on_event('shutdown')
async def shutdown():
    mongo_client.close()
